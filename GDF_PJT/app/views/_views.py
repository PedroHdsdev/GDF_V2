"""
Views do app GDF – ponto único: telas e APIs.
- Telas (fn_view_*): login, home, CRUD usuários/empresas/clientes, importação (Carga XML, Carga SPED), Reprocessamento.
- APIs (fn_api_*): CargaXml, CargaSped, Relatórios, Reprocessamento, SAP, Sessão.
- Usa app.classes (ClGdf, CargaXml, CargaSped) e app.utils.view_helpers. Jobs em app.api.jobs.
"""
import json
import math
import os
import re
import unicodedata
import shutil
import tempfile
import threading
import zipfile
from datetime import datetime, timedelta, timezone as _py_tz
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt

from app.classes.CargaSped import CargaSped
from app.classes.CargaXml import CargaXml
from app.classes.gdf import ClGdf
from app.classes.Reprocessamento import (
    _condicao_sap_da_param,
    condicao_pagamento_da_nfe,
    tipo_pagamento_da_nfe,
)
from app.db_GDF.Public.models import (
    AcessoSubsolucaoGrupo,
    ClienteGdf,
    ConexaoSap,
    Empresa,
    Filial,
    JobCargaSped,
    JobCargaXml,
    UsuarioEmpresa,
)
from app.db_GDF.CTe.models import (
    CTe,
    CTe_Carga,
    CTe_Destinatario,
    CTe_Emitente,
    CTe_Fiscal,
    CTe_Identificacao,
    CTe_Motorista,
    CTe_Percurso,
    CTe_Servico,
    CTe_Transporte,
    CTe_Valor,
    CTe_Veiculo,
)
from app.db_GDF.NFe.models import (
    NFe,
    NFe_Cobranca,
    NFe_Destinatario,
    NFe_Emitente,
    NFe_Endereco,
    NFe_Identificacao,
    NFe_Informacoes_Adicionais,
    NFe_Pagamento,
    NFe_Parcela,
    NFe_Produto,
    NFe_Total,
    NFe_Transporte,
)
from app.db_GDF.NFSe.models import (
    NFSe,
    NFSe_Endereco,
    NFSe_Identificacao,
    NFSe_Pagamento,
    NFSe_Prestador,
    NFSe_Retencao,
    NFSe_RPS,
    NFSe_Servico,
    NFSe_Tomador,
)
from app.db_GDF.sped_fiscal.models import SpedFiscalArquivo, SpedFiscalReg_C100, SpedFiscalReg_C170
from app.db_GDF.reprocessamento.models import (
    CondicaoPagamentoLote,
    CondicaoParam,
    Divergencia,
    ReprocessamentoJob,
    ReprocessamentoLote,
)
from app.security.decorators import (
    validate_idor_empresa,
    validate_idor_usuario,
    validate_session_required,
    requer_acesso_subsolucao,
    requer_acesso_total_painel,
    requer_superuser,
)
from app.security.validators import InputValidator
from app.security_logger import SecurityLogger
from django.core.exceptions import ValidationError
from app.utils.view_helpers import (
    COD_CLIENTE_PROJETO,
    TIPO_PAGAMENTO_DESC,
    autenticar_sessao_ou_jwt_dashboard,
    descricao_tipo_pagamento,
    get_subsolucoes_usuario,
    relatorio_empresas_queryset,
    reprocessamento_empresas_cliente,
    superuser_acesso_total_painel,
    usuario_acesso_total_painel,
    usuario_vinculado_cliente_1000,
)
from app.utils.relatorio_params import (
    paginate_queryset,
    parse_date_safe,
    parse_filial_id,
    parse_relatorio_order,
    parse_relatorio_params,
)
from app.utils.relatorio_querysets import (
    list_relatorio_sped_items,
    queryset_relatorio_cte,
    queryset_relatorio_nfe,
    queryset_relatorio_nfse,
)
from app.utils.datetime_json import isoformat_brasilia

@ensure_csrf_cookie
def fn_view_login(request):
    if request.method == "POST":
        Username = request.POST.get('Username')
        password = request.POST.get('password') 

        user = authenticate(username=Username, password=password)
        
        if user is not None:
            if not getattr(user, 'is_active', True):
                SecurityLogger.log_login_attempt(request, False, reason='Usuário inativo')
                return render(request, 'comum/login.html', {'error_message': 'Usuário inativo.'})
            login(request, user)
            SecurityLogger.log_login_attempt(request, True)
            cl_gdf_instance = ClGdf()
            cl_gdf_instance.get_dados(request.user)

            request.session['is_superuser'] = getattr(user, 'is_superuser', False)
            request.session['is_staff'] = getattr(user, 'is_staff', False)
            # Superuser vinculado ao cliente dono do projeto (PRCIT): acesso total (compatibilidade)
            _cliente = getattr(cl_gdf_instance, 'Cliente', None)
            _cod = getattr(_cliente, 'cod_cliente', None) if _cliente else None
            request.session['superuser_cliente_1000'] = (
                getattr(user, 'is_superuser', False) and _cod is not None and str(_cod).strip() == COD_CLIENTE_PROJETO
            )
            # Usuário (não superuser) vinculado ao cliente PRCIT: acesso total como empresa dona do projeto
            request.session['usuario_cliente_1000'] = (
                not getattr(user, 'is_superuser', False) and usuario_vinculado_cliente_1000(user)
            )

            if not cl_gdf_instance.Retorn:
                solucoes = cl_gdf_instance.get_solucoes()
                cod_cliente = (
                    cl_gdf_instance.ClienteGdf.cod_cliente
                    if getattr(cl_gdf_instance, 'ClienteGdf', None) else None
                )
                cod_cliente = (cod_cliente or '').strip() or None
                # Primeira vez / sem cliente: superuser ou cliente PRCIT começam com cliente dono do projeto
                if not cod_cliente and (getattr(user, 'is_superuser', False) or request.session.get('usuario_cliente_1000', False)):
                    cod_cliente = COD_CLIENTE_PROJETO
                # Permitir login se tiver soluções OU se tiver ao menos cliente (empresas vinculadas), mesmo sem subsoluções nos grupos
                if solucoes or getattr(user, 'is_superuser', False):
                    request.session['t_solucoes'] = solucoes or []
                    request.session['cod_cliente'] = cod_cliente
                    return redirect('Home')
                if cod_cliente:
                    # Usuário tem cliente/empresas mas grupos sem subsoluções liberadas: permite login com menu vazio
                    request.session['t_solucoes'] = []
                    request.session['cod_cliente'] = cod_cliente
                    return redirect('Home')
                SecurityLogger.log_login_attempt(request, False, reason='Problema de Acesso (sem cliente/subsoluções)')
                return render(request, 'comum/login.html', {
                    'error_message': 'Problema de Acesso. Garanta que: (1) as empresas do usuário tenham um cliente vinculado (campo Cliente na empresa) ou que os grupos do usuário estejam vinculados a um cliente (Permissão grupo-cliente); (2) o cliente tenha soluções ativas (AcessoSolucaoCliente); (3) os grupos tenham subsoluções (AcessoSubsolucaoGrupo) no Admin.'
                })
            # Redirecionamento sem solucoes (ex.: Retorn True): manter cliente dono do projeto (PRCIT) como padrão
            if not request.session.get('cod_cliente') and (getattr(user, 'is_superuser', False) or request.session.get('usuario_cliente_1000', False)):
                request.session['cod_cliente'] = COD_CLIENTE_PROJETO
            return redirect('Home')
        else:
            SecurityLogger.log_login_attempt(request, False, reason='Usuário ou senha inválidos')
            return render(request, 'comum/login.html', {'error_message': 'Usuário ou senha inválidos.'})

    return render(request, 'comum/login.html')


def fn_view_csrf_failure(request, reason=''):
    """Redireciona para a tela de login quando o token CSRF falha (ex.: voltar após login)."""
    from django.urls import reverse
    return redirect(reverse('Login'))


@login_required(login_url='Login')
def fn_view_obter_subsolucao(request, cod_sub):
    """Redireciona para a URL da subsolução apenas se o usuário tem acesso via grupo."""
    if str(cod_sub) == 'Dm_Filiais':
        s = get_subsolucoes_usuario(request.user)
        ok = s is None or 'Dm_Filiais' in s or 'Dm_Empresas' in s
        if not ok:
            return redirect('Home')
        return redirect('Dm_Empresas')
    # Valida acesso pela fonte de verdade (grupos), não apenas pela sessão
    subsolucoes = get_subsolucoes_usuario(request.user)
    if subsolucoes is not None and str(cod_sub) not in {str(c) for c in subsolucoes}:
        return redirect('Home')
    solucoes = request.session.get('t_solucoes', [])
    for sol in solucoes:
        for sub in sol.get('sub_solucoes', []):
            if str(sub.get('cod_subsolucao')) == str(cod_sub):
                return redirect(sub.get('cod_subsolucao'))
    return render(request, 'home/inicio.html')

@login_required(login_url='Login')
def fn_view_home(request):
    if not request.user.is_authenticated:
        return redirect('Login')
    cod_cliente = request.session.get('cod_cliente')
    # Superuser ou usuário cliente dono do projeto (PRCIT): permitir trocar cliente por POST
    _REDIRECT_NAMES = ('Home', 'Dm_Empresas', 'Dm_Usuarios', 'Dm_Clientes', 'Dm_Filiais')
    if request.method == "POST":
        codigo = request.POST.get('codigo')
        novo_cliente = request.POST.get('cod_cliente', '').strip()
        next_name = (request.POST.get('next') or '').strip()
        if usuario_acesso_total_painel(request) and novo_cliente:
            request.session['cod_cliente'] = novo_cliente
            if next_name in _REDIRECT_NAMES:
                return redirect(next_name)
            if next_name.startswith('/'):
                return redirect(next_name)
            return redirect('Home')
        if codigo:
            return redirect(codigo)
    context = {'cod_cliente': cod_cliente}
    if usuario_acesso_total_painel(request):
        context['is_superuser'] = request.session.get('is_superuser', False)
        cl_gdf = ClGdf()
        context['lista_clientes'] = cl_gdf.get_clientes()

    # Subsoluções que o usuário tem acesso (None = superuser = acesso total)
    subsolucoes = get_subsolucoes_usuario(request.user)

    def _tem_acesso(cod):
        """Verifica se usuário tem acesso à subsolução."""
        if subsolucoes is None:
            return True
        return cod in subsolucoes

    # Dados reais para alertas e métricas (evita dados falsos)
    context['alertas'] = []
    context['metricas'] = {'cert_expirando': 0, 'carga_xml_24h': 0, 'carga_sped_24h': 0, 'carga_em_andamento': 0}
    desde_24h = timezone.now() - timedelta(hours=24)

    if cod_cliente:
        hoje = timezone.now().date()
        limite_cert = hoje + timedelta(days=30)

        # 1. Certificados expirando (requer Dm_Empresas)
        if _tem_acesso('Dm_Empresas'):
            cert_expirando = Empresa.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                cert__isnull=False,
                cert__fim_validade__isnull=False,
            ).filter(cert__fim_validade__date__lte=limite_cert).count()
            context['metricas']['cert_expirando'] = cert_expirando
            if cert_expirando > 0:
                context['alertas'].append({
                    'tipo': 'warning',
                    'titulo': f'{cert_expirando} certificado(s) expirando em até 30 dias',
                    'meta': 'Empresas',
                    'tag': 'Atenção',
                    'url': 'Dm_Empresas',
                })

        # 2. Carga XML com erros (requer Pro_CargaXml) – ids para o cliente esconder se "já lido"
        if _tem_acesso('Pro_CargaXml'):
            xml_erros_qs = JobCargaXml.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status='ERROR',
                finished_at__gte=desde_24h,
            )
            xml_erro_ids = list(xml_erros_qs.values_list('id', flat=True))
            if xml_erro_ids:
                context['alertas'].append({
                    'tipo': 'critical',
                    'titulo': f'{len(xml_erro_ids)} carga(s) XML com erro nas últimas 24h',
                    'meta': 'Importação',
                    'tag': 'Urgente',
                    'url': 'Pro_CargaXml',
                    'ids': xml_erro_ids,
                    'fonte': 'cargaxml',
                })

        # 3. Carga SPED com erros (requer Pro_CargaSped) – ids para o cliente esconder se "já lido"
        if _tem_acesso('Pro_CargaSped'):
            sped_erros_qs = JobCargaSped.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status='ERROR',
                finished_at__gte=desde_24h,
            )
            sped_erro_ids = list(sped_erros_qs.values_list('id', flat=True))
            if sped_erro_ids:
                context['alertas'].append({
                    'tipo': 'critical',
                    'titulo': f'{len(sped_erro_ids)} carga(s) SPED com erro nas últimas 24h',
                    'meta': 'Importação',
                    'tag': 'Urgente',
                    'url': 'Pro_CargaSped',
                    'ids': sped_erro_ids,
                    'fonte': 'cargasped',
                })

        # 4. Divergências abertas no painel de reprocessamento (requer Reproc_Painel)
        if _tem_acesso('Reproc_Painel'):
            divergencias = Divergencia.objects.filter(
                lote__empresa__gdfcliente_id=cod_cliente,
                status='ABERTA',
            ).count()
            if divergencias > 0:
                context['alertas'].append({
                    'tipo': 'warning',
                    'titulo': f'{divergencias} divergência(s) aberta(s) no confronto SPED x NFe',
                    'meta': 'Ferramentas',
                    'tag': 'Revisar',
                    'url': 'Reproc_Painel',
                })

        # 5. Atalhos informativos (apenas um por área, para quem tem acesso)
        if _tem_acesso('Pro_CargaXml') or _tem_acesso('Pro_CargaSped'):
            context['alertas'].append({
                'tipo': 'info',
                'titulo': 'Carga fiscal disponível',
                'meta': 'XML e SPED • Importar documentos',
                'tag': 'Pronto',
                'url': 'Pro_CargaXml' if _tem_acesso('Pro_CargaXml') else 'Pro_CargaSped',
            })
        if _tem_acesso('Mnf_Painel'):
            context['alertas'].append({
                'tipo': 'info',
                'titulo': 'Painel manifesto',
                'meta': 'NFe, CTe e NFSe',
                'tag': 'Acessar',
                'url': 'Mnf_Painel',
            })
        if _tem_acesso('Reproc_Painel'):
            context['alertas'].append({
                'tipo': 'info',
                'titulo': 'Confronto SPED x NFe',
                'meta': 'Ferramentas',
                'tag': 'Acessar',
                'url': 'Reproc_Painel',
            })

        # Métricas (apenas para quem tem acesso às cargas)
        if _tem_acesso('Pro_CargaXml') or _tem_acesso('Pro_CargaSped'):
            xml_concluidos = JobCargaXml.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status='SUCCESS',
                finished_at__gte=desde_24h,
            ).count() if _tem_acesso('Pro_CargaXml') else 0
            sped_concluidos = JobCargaSped.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status='SUCCESS',
                finished_at__gte=desde_24h,
            ).count() if _tem_acesso('Pro_CargaSped') else 0
            xml_em_andamento = JobCargaXml.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status__in=('RUNNING', 'PENDING'),
            ).count() if _tem_acesso('Pro_CargaXml') else 0
            sped_em_andamento = JobCargaSped.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
                status__in=('RUNNING', 'PENDING'),
            ).count() if _tem_acesso('Pro_CargaSped') else 0
            context['metricas'].update({
                'carga_xml_24h': xml_concluidos,
                'carga_sped_24h': sped_concluidos,
                'carga_em_andamento': xml_em_andamento + sped_em_andamento,
            })

    # Se não há alertas (ex.: sem cliente ou sem permissões), mensagem neutra
    if not context['alertas']:
        context['alertas'].append({
            'tipo': 'info',
            'titulo': 'Nenhum alerta no momento',
            'meta': 'Selecione um cliente ou verifique suas permissões',
            'tag': 'OK',
            'url': None,
        })

    # Atalhos filtrados por permissão (para o card Acesso rápido)
    context['atalhos'] = []
    context['tem_mnf'] = _tem_acesso('Mnf_Painel')
    context['tem_empresas'] = _tem_acesso('Dm_Empresas')
    _atalhos_config = [
        ('Pro_CargaXml', 'Importar XML', 'Carga e relatórios NFe, CT-e, NFS'),
        ('Pro_CargaSped', 'Carga SPED', 'Arquivos e relatório SPED'),
        ('Int_Rfc', 'RFC SAP', 'Ferramentas · integração schema SAP'),
        ('Dm_Empresas', 'Empresas (e filiais)', 'Cadastros'),
        ('Dm_Usuarios', 'Usuários', 'Acessos'),
    ]
    for cod, titulo, desc in _atalhos_config:
        if _tem_acesso(cod):
            context['atalhos'].append({'url': cod, 'titulo': titulo, 'desc': desc})

    # Contexto enriquecido: welcome, cliente, competência, estatísticas, atividade recente
    context['nome_usuario'] = request.user.get_full_name() or request.user.username
    _meses = ('', 'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
              'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro')
    _agora = timezone.now()
    context['competencia'] = f"{_meses[_agora.month]}/{_agora.year}"
    context['stats_docs'] = {}
    context['ultima_atividade'] = []
    context['cliente_nome'] = None
    context['qtd_empresas'] = 0

    if cod_cliente:
        cliente_obj = ClienteGdf.objects.filter(cod_cliente=cod_cliente).first()
        if cliente_obj:
            context['cliente_nome'] = cliente_obj.razao or cod_cliente
        context['qtd_empresas'] = Empresa.objects.filter(gdfcliente__cod_cliente=cod_cliente).count()

        # Documentos no mês atual (importação XML/SPED ou reprocessamento)
        if _tem_acesso('Pro_CargaXml') or _tem_acesso('Pro_CargaSped') or _tem_acesso('Reproc_Painel'):
            hoje = timezone.now()
            nfe_mes = NFe.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
            #    identificacao__emissao__year=hoje.year,
            #    identificacao__emissao__month=hoje.month,
            ).count()
            cte_mes = CTe.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
            #    identificacao__emissao__year=hoje.year,
            #    identificacao__emissao__month=hoje.month,
            ).count()
            nfse_mes = NFSe.objects.filter(
                gdfcliente__cod_cliente=cod_cliente,
            #    identificacao__emissao__year=hoje.year,
            #    identificacao__emissao__month=hoje.month,
            ).count()
            
            nfe_mes_fmt  = ClGdf.formatar_numero(nfe_mes)
            cte_mes_fmt  = ClGdf.formatar_numero(cte_mes)
            nfse_mes_fmt = ClGdf.formatar_numero(nfse_mes)
            
            context['stats_docs'] = {'nfe': nfe_mes_fmt, 'cte': cte_mes_fmt, 'nfse': nfse_mes_fmt}

        # Última atividade: jobs recentes (XML + SPED)
        if _tem_acesso('Pro_CargaXml') or _tem_acesso('Pro_CargaSped'):
            atividades = []
            if _tem_acesso('Pro_CargaXml'):
                for j in JobCargaXml.objects.filter(gdfcliente__cod_cliente=cod_cliente).order_by('-started_at')[:3]:
                    dt = j.finished_at or j.started_at
                    atividades.append({
                        'tipo': 'XML',
                        'status': j.status,
                        'data': dt,
                        'total': j.total_arquivos,
                        'sucesso': j.total_sucesso,
                        'erro': j.total_erro,
                        'url': 'Pro_CargaXml',
                    })
            if _tem_acesso('Pro_CargaSped'):
                for j in JobCargaSped.objects.filter(gdfcliente__cod_cliente=cod_cliente).order_by('-started_at')[:3]:
                    dt = j.finished_at or j.started_at
                    atividades.append({
                        'tipo': 'SPED',
                        'status': j.status,
                        'data': dt,
                        'total': j.total_arquivos,
                        'sucesso': j.total_sucesso,
                        'erro': j.total_erro,
                        'url': 'Pro_CargaSped',
                    })
            # Ordenar por data e pegar os 5 mais recentes (None vai por último)
            _epoch = datetime(1970, 1, 1, tzinfo=_py_tz.utc)
            atividades.sort(key=lambda x: x['data'] or _epoch, reverse=True)
            context['ultima_atividade'] = atividades[:5]

    return render(request, "home/inicio.html", context)

@login_required
def fn_view_sair(request):   
    logout(request)
    return redirect('Login')

#--------------------------------------------------------------------
#       Sub-soluções Views (Configuração)
#--------------------------------------------------------------------
# Usuarios
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Usuarios')
def fn_view_listar_usuarios(request):
    cod_cliente = request.session.get('cod_cliente', None)
    is_superuser = request.session.get('is_superuser', False)
    if not cod_cliente:
        if is_superuser:
            messages.info(request, 'Selecione um cliente na Home para gerenciar usuários.')
            return redirect('Home')
        return render(request, 'comum/login.html', {'error_message': 'Acesso negado: cliente não identificado'})
    
    cl_gdf = ClGdf()
    t_user = cl_gdf.get_usuarios(i_v_cod_cliente=cod_cliente)
    is_superuser = request.session.get('is_superuser', False)
    context = {
        't_user': t_user,
        'cod_cliente': cod_cliente,
        'is_superuser': is_superuser,
    }
    if is_superuser and superuser_acesso_total_painel(request):
        context['lista_clientes'] = cl_gdf.get_clientes()
    return render(request, 'usuarios/index.html', context)

# Empresas
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
def fn_view_listar_empresas(request):
    cod_cliente = request.session.get('cod_cliente', None)
    is_superuser = request.session.get('is_superuser', False)
    if not cod_cliente:
        if is_superuser:
            messages.info(request, 'Selecione um cliente na Home para gerenciar empresas.')
            return redirect('Home')
        return redirect('Login')
    
    cl_gdf = ClGdf()
    t_empresas = cl_gdf.get_empresas(i_v_cod_cliente=cod_cliente)
    context = {
        't_empresas': t_empresas,
        'cod_cliente': cod_cliente,
        'is_superuser': is_superuser,
        # Filiais: só no modal de empresa; quem acessa esta tela já tem Dm_Empresas
        'pode_gerir_filiais': True,
    }
    if is_superuser and superuser_acesso_total_painel(request):
        context['lista_clientes'] = cl_gdf.get_clientes()
    return render(request, 'empresas/index.html', context)

# Clientes
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
def fn_view_listar_clientes(request):
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente and not usuario_acesso_total_painel(request):
        return redirect('Login')
    cl_gdf = ClGdf()
    t_clientes = cl_gdf.get_clientes()
    context = {'t_clientes': t_clientes, 'cod_cliente': cod_cliente}
    if usuario_acesso_total_painel(request):
        context['is_superuser'] = request.session.get('is_superuser', False)
    return render(request, 'mandantes/index.html', context)


# Rota Dm_Filiais mantida por compat.; filiais passam a ser gestas só no modal de Empresas
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
def fn_view_listar_filiais(request):
    messages.info(
        request,
        'Filiais: use Cadastros → Empresas, clique na linha da empresa e aba Filiais no modal.',
    )
    return redirect('Dm_Empresas')


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
@require_http_methods(["POST"])
def fn_view_inserir_filial(request):
    """Cadastrar nova filial (empresa do cliente na sessão)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    cod_empresa = request.POST.get("m_empresa", "").strip()
    cod_filial = request.POST.get("m_cod_filial", "").strip()
    nome = request.POST.get("m_nome", "").strip()
    cnpj = request.POST.get("m_cnpj", "").strip()
    ativo = request.POST.get("m_ativo") == "on" or request.POST.get("m_ativo") == "true"
    errors = []
    if not cod_empresa:
        errors.append("Empresa é obrigatória")
    if not cod_filial:
        errors.append("Código da filial é obrigatório")
    if errors:
        return JsonResponse({"erro": " | ".join(errors)}, status=400)
    try:
        empresa = Empresa.objects.get(
            cod_empresa=cod_empresa,
            gdfcliente_id=cod_cliente,
        )
    except Empresa.DoesNotExist:
        return JsonResponse({"erro": "Empresa não encontrada ou não pertence ao cliente"}, status=400)
    if Filial.objects.filter(empresa=empresa, cod_filial=cod_filial).exists():
        return JsonResponse({"erro": f"Já existe filial com código '{cod_filial}' para esta empresa"}, status=400)
    Filial.objects.create(
        empresa=empresa,
        cod_filial=cod_filial,
        nome=nome or None,
        cnpj=cnpj.replace(" ", "").replace(".", "").replace("/", "").replace("-", "") or None,
        ativo=ativo,
    )
    return JsonResponse({"success": True, "message": "Filial cadastrada com sucesso"})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
@require_http_methods(["GET", "POST"])
def fn_view_atualizar_filial(request, pk):
    """Retorna dados da filial (GET) ou atualiza (POST). Filial deve pertencer ao cliente da sessão."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    filial = (
        Filial.objects.filter(
            id=pk,
            empresa__gdfcliente__cod_cliente=cod_cliente,
        )
        .select_related('empresa')
        .first()
    )
    if not filial:
        return JsonResponse({"erro": "Filial não encontrada ou sem permissão"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": filial.id,
            "cod_filial": filial.cod_filial,
            "nome": filial.nome or "",
            "cnpj": filial.cnpj or "",
            "ativo": filial.ativo,
            "empresa_cod": filial.empresa.cod_empresa,
            "empresa_nome": filial.empresa.fantasia or filial.empresa.razao,
        })

    # POST: atualizar
    cod_filial = request.POST.get("m_cod_filial", "").strip()
    nome = request.POST.get("m_nome", "").strip()
    cnpj = request.POST.get("m_cnpj", "").strip()
    ativo = request.POST.get("m_ativo") in ("on", "true", "1")
    if not cod_filial:
        return JsonResponse({"erro": "Código da filial é obrigatório"}, status=400)
    if Filial.objects.filter(empresa=filial.empresa, cod_filial=cod_filial).exclude(id=filial.id).exists():
        return JsonResponse({"erro": f"Já existe outra filial com código '{cod_filial}' nesta empresa"}, status=400)
    cnpj_limpo = (cnpj or "").replace(" ", "").replace(".", "").replace("/", "").replace("-", "").strip() or None
    filial.cod_filial = cod_filial
    filial.nome = nome or None
    filial.cnpj = cnpj_limpo
    filial.ativo = ativo
    filial.save(update_fields=["cod_filial", "nome", "cnpj", "ativo"])
    return JsonResponse({"success": True, "message": "Filial atualizada com sucesso"})


@login_required(login_url="Login")
@requer_acesso_subsolucao("Dm_Empresas")
@require_http_methods(["GET"])
def fn_view_listar_filiais_empresa(request, cod_empresa):
    """Lista filiais (JSON) da empresa, desde que pertença ao cliente da sessão."""
    cod_cliente = request.session.get("cod_cliente", None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    empresa = Empresa.objects.filter(
        cod_empresa=cod_empresa,
        gdfcliente__cod_cliente=cod_cliente,
    ).first()
    if not empresa:
        return JsonResponse({"erro": "Empresa não encontrada ou sem permissão"}, status=404)
    filiais = (
        Filial.objects.filter(empresa=empresa)
        .order_by("cod_filial")
    )
    data = [
        {
            "id": f.id,
            "cod_filial": f.cod_filial,
            "nome": f.nome or "",
            "cnpj": f.cnpj or "",
            "ativo": f.ativo,
        }
        for f in filiais
    ]
    return JsonResponse({"filiais": data})


@login_required(login_url="Login")
@requer_acesso_subsolucao("Dm_Empresas")
@require_http_methods(["POST"])
def fn_view_excluir_filial(request, pk):
    """Exclui filial do cliente (POST). Pode falhar com PROTECT (ex.: SAP)."""
    cod_cliente = request.session.get("cod_cliente", None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    filial = (
        Filial.objects.filter(
            id=pk,
            empresa__gdfcliente__cod_cliente=cod_cliente,
        )
        .select_related("empresa")
        .first()
    )
    if not filial:
        return JsonResponse({"erro": "Filial não encontrada ou sem permissão"}, status=404)
    try:
        filial.delete()
    except ProtectedError:
        return JsonResponse(
            {
                "erro": "Não é possível excluir: existem registros vinculados a esta filial (ex.: integração SAP).",
            },
            status=400,
        )
    return JsonResponse({"success": True, "message": "Filial excluída com sucesso."})


#--------------------------------------------------------------------
#       Modais Views
#--------------------------------------------------------------------
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Usuarios')
@require_http_methods(["GET", "POST"])
def fn_view_inserir_usuario(request):
    """Inserir usuário. Superuser pode informar o cliente no formulário."""
    is_superuser = request.session.get('is_superuser', False)
    cod_cliente = request.session.get('cod_cliente', None)
    if request.method == "GET":
        if is_superuser and request.GET.get('cod_cliente'):
            cod_cliente = request.GET.get('cod_cliente', '').strip() or cod_cliente
        if not cod_cliente:
            return JsonResponse({"erro": "Cliente não identificado"}, status=403)
        cl_gdf = ClGdf()
        return JsonResponse(cl_gdf.get_usuario_dados_ins(i_v_cod_cliente=cod_cliente))

    if request.method == "POST":
        if is_superuser and request.POST.get("m_cod_cliente"):
            cod_cliente = request.POST.get("m_cod_cliente", "").strip() or cod_cliente
        if not cod_cliente:
            return JsonResponse({"erro": "Cliente não identificado"}, status=403)
        cl_gdf = ClGdf()
        username        = request.POST.get("username", "").strip()
        first_name      = request.POST.get("first_name", "").strip()
        last_name       = request.POST.get("last_name", "").strip()
        email           = request.POST.get("email", "").strip()
        password        = request.POST.get("password", "").strip()
        password_conf   = request.POST.get("password_confirm", "").strip()
        empresas_str    = request.POST.get("ls_empresas", "").strip()
        grupos_str      = request.POST.get("ls_grupos", "").strip()

        errors = []
        if not username:
            errors.append("Username é obrigatório")
        if not email:
            errors.append("Email é obrigatório")
        if not password:
            errors.append("Senha é obrigatória")
        if password != password_conf:
            errors.append("Senhas não conferem")
        if not empresas_str:
            errors.append("Selecione pelo menos 1 empresa")
        if not grupos_str:
            errors.append("Selecione pelo menos 1 grupo")
        if errors:
            t_user = cl_gdf.get_usuarios(i_v_cod_cliente=cod_cliente)
            ctx = {'t_user': t_user, 'error_message': ' | '.join(errors), 'cod_cliente': cod_cliente, 'is_superuser': is_superuser}
            if is_superuser and superuser_acesso_total_painel(request):
                ctx['lista_clientes'] = cl_gdf.get_clientes()
            return render(request, 'usuarios/index.html', ctx)

        resultado = cl_gdf.set_usuario(
            i_v_username=username,
            i_v_email=email,
            i_v_password=password,
            i_v_first_name=first_name,
            i_v_last_name=last_name,
            i_lsl_empresas_ids=empresas_str,
            i_lsl_grupos_ids=grupos_str,
            i_v_cod_cliente=cod_cliente
        )
        if not resultado.get("success"):
            t_user = cl_gdf.get_usuarios(i_v_cod_cliente=cod_cliente)
            ctx = {'t_user': t_user, 'error_message': resultado.get("message", "Erro ao criar usuário"), 'cod_cliente': cod_cliente, 'is_superuser': is_superuser}
            if is_superuser and superuser_acesso_total_painel(request):
                ctx['lista_clientes'] = cl_gdf.get_clientes()
            return render(request, 'usuarios/index.html', ctx)
        return redirect('Dm_Usuarios')

@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Usuarios')
@require_http_methods(["GET", "POST"])
def fn_view_atualizar_usuario(request, user_id):
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    
    # ✅ VALIDAR IDOR: User só pode editar usuários de suas empresas
    user_belongs_to_client = UsuarioEmpresa.objects.filter(
        user_id=user_id,
        empresa__gdfcliente__cod_cliente=cod_cliente
    ).exists()
    
    if not user_belongs_to_client:
        return JsonResponse({"erro": "Acesso negado: usuário não pertence ao seu cliente"}, status=403)
    
    cl_gdf = ClGdf()
    if request.method == "GET":
        # Retornar dados do usuário em JSON (para modal)
        data = cl_gdf.get_usuario_upd(i_v_user_id=int(user_id), i_v_cod_cliente=cod_cliente)
        if not data or data.get('erro'):
            return JsonResponse({"erro": "Usuário não encontrado"}, status=404)

        return JsonResponse(data)

    elif request.method == "POST":
        # ✅ Validações de campos obrigatórios
        first_name  = request.POST.get("upd_first_name", "").strip()
        last_name   = request.POST.get("upd_last_name", "").strip()
        email       = request.POST.get("upd_email", "").strip()
        is_active   = request.POST.get("upd_is_active") == "on"

        empresas_str = request.POST.get("ls_empresas", "")
        empresa_ids = [e.strip() for e in empresas_str.split(",") if e.strip()]

        # ✅ Processar grupos: podem vir como string separada por vírgula do hidden input
        grupos_str = request.POST.get("ls_grupos", "")
        try:
            grupo_ids = [int(g.strip()) for g in grupos_str.split(",") if g.strip()]
        except ValueError:
            return JsonResponse({"erro": "IDs de grupos inválidos (devem ser numéricos)"}, status=400)

        # ✅ Validações básicas antes de chamar método
        if not email:
            return JsonResponse({"erro": "Email obrigatório"}, status=400)
        if not empresa_ids:
            return JsonResponse({"erro": "Selecione pelo menos 1 empresa"}, status=400)
        if not grupo_ids:
            return JsonResponse({"erro": "Selecione pelo menos 1 grupo"}, status=400)

        new_password = request.POST.get("upd_new_password", "").strip()
        new_password_confirm = request.POST.get("upd_new_password_confirm", "").strip()
        if new_password or new_password_confirm:
            if new_password != new_password_confirm:
                return JsonResponse({"erro": "Nova senha e confirmação não conferem"}, status=400)
            if len(new_password) < 6:
                return JsonResponse({"erro": "A nova senha deve ter no mínimo 6 caracteres"}, status=400)

        resultado = cl_gdf.upd_usuario(
            i_v_user_id=int(user_id),
            i_v_first_name=first_name,
            i_v_last_name=last_name,
            i_v_email=email,
            i_v_is_active=is_active,
            i_lsl_empresa_ids=empresa_ids,
            i_lsl_grupo_ids=grupo_ids,
            i_v_cod_cliente=cod_cliente,
            i_v_new_password=new_password or None,
        )
        
        # ✅ Detectar se é requisição AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # ✅ Verificar resultado
        if not resultado.get("success"):
            mensagem = resultado.get("message", "Erro ao atualizar")
            if is_ajax:
                return JsonResponse({"success": False, "message": mensagem}, status=400)
            else:
                messages.error(request, mensagem, extra_tags='MODAL_UPD')
                return redirect('Dm_Usuarios')
        
        mensagem = resultado.get("message", "Usuário atualizado com sucesso!")
        if is_ajax:
            return JsonResponse({"success": True, "message": mensagem}, status=200)
        else:
            messages.success(request, mensagem, extra_tags='MODAL_UPD')
            return redirect('Dm_Usuarios')
    
    return JsonResponse({"erro": "Método não permitido"}, status=405)
    
#--------------------------------------------------------------------
#       Sub-soluções Views (Dashboard)
#--------------------------------------------------------------------
def _streamlit_iframe_url(request):
    """URL absoluta do Streamlit para o iframe; se a config for relativa, monta a partir do request."""
    streamlit_url = getattr(settings, 'STREAMLIT_IFRAME_URL', 'http://127.0.0.1:8600').rstrip('/')
    if streamlit_url.startswith(('http://', 'https://')):
        return streamlit_url
    base = request.build_absolute_uri('/').rstrip('/')
    path = (streamlit_url or 'streamlit').strip().lstrip('/') or 'streamlit'
    return f"{base}/{path}"


@login_required(login_url='Login')
@requer_acesso_subsolucao('Db_Vendas')
def fn_view_dashboard_vendas(request):
    token = ClGdf.gerar_token(request, request.user, tipo_relatorio='Vendas')
    if not token:
        return render(request, 'comum/login.html', {'error_message': 'Erro ao gerar token de acesso'})
    streamlit_url = _streamlit_iframe_url(request)
    return render(request, "dashboard/vendas.html", {"token": token, "streamlit_iframe_url": streamlit_url})

@login_required(login_url='Login')
@requer_acesso_subsolucao('Db_Compras')
def fn_view_dashboard_compras(request):
    token = ClGdf.gerar_token(request, request.user, tipo_relatorio='Compras')
    if not token:
        return render(request, 'comum/login.html', {'error_message': 'Erro ao gerar token de acesso'})
    streamlit_url = _streamlit_iframe_url(request)
    return render(request, "dashboard/compras.html", {"token": token, "streamlit_iframe_url": streamlit_url})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Db_Custo')
def fn_view_dashboard_custo(request):
    token = ClGdf.gerar_token(request, request.user, tipo_relatorio='Custo')
    if not token:
        return render(request, 'comum/login.html', {'error_message': 'Erro ao gerar token de acesso'})
    streamlit_url = _streamlit_iframe_url(request)
    return render(request, "dashboard/custo.html", {"token": token, "streamlit_iframe_url": streamlit_url})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Db_DemonstrContabeis')
def fn_view_dashboard_demonstrativos_contabeis(request):
    """Dashboard Demonstrativos contábeis (Streamlit); token tipo_relatorio DemonstrativosContabeis."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return render(request, 'comum/login.html', {'error_message': 'Cliente não identificado'})
    token = ClGdf.gerar_token(request, request.user, tipo_relatorio='DemonstrativosContabeis')
    if not token:
        return render(request, 'comum/login.html', {'error_message': 'Erro ao gerar token de acesso'})
    streamlit_url = _streamlit_iframe_url(request)
    return render(
        request,
        "dashboard/demonstrativos_contabeis.html",
        {"token": token, "streamlit_iframe_url": streamlit_url},
    )


#--------------------------------------------------------------------
#       Sub-soluções Views (Manifesto)
#--------------------------------------------------------------------
@login_required(login_url='Login')
@requer_acesso_subsolucao('Mnf_Painel')
def fn_view_manifesto_painel(request):
    manifesto_data = {
        "notas": [
            {
                "id": "NFE-2026-0001",
                "tipo": "NFe",
                "numero": "15432",
                "serie": "001",
                "emissao": "2026-02-10 09:22",
                "emitente": "Industrias Atlas",
                "destinatario": "Supermercado Central",
                "valor": "152340.80",
                "status": "AUTORIZADA",
                "documentos": [
                    {
                        "tipo": "COMPRA",
                        "numero": "4500012391",
                        "data": "2026-02-10",
                        "status": "CRIADO",
                        "itens": [
                            {"seq": 10, "nfe_item_seq": 1, "material": "MAT-1001", "descricao": "Acucar cristal", "qtd": "100", "un": "KG", "valor": "3400.00"},
                            {"seq": 20, "nfe_item_seq": 2, "material": "MAT-2004", "descricao": "Leite integral", "qtd": "220", "un": "CX", "valor": "6600.00"},
                        ],
                    },
                    {
                        "tipo": "MIRO",
                        "numero": "5100004890",
                        "data": "2026-02-10",
                        "status": "PENDENTE",
                        "itens": [
                            {"seq": 10, "nfe_item_seq": 1, "material": "MAT-1001", "descricao": "Acucar cristal", "qtd": "100", "un": "KG", "valor": "3400.00"},
                        ],
                    },
                ],
                "itens": [
                    {"seq": 1, "codigo": "7891001", "descricao": "Acucar cristal 5kg", "qtd": "100", "un": "KG", "valor": "3400.00"},
                    {"seq": 2, "codigo": "7892004", "descricao": "Leite integral 1L", "qtd": "220", "un": "CX", "valor": "6600.00"},
                    {"seq": 3, "codigo": "7893011", "descricao": "Cafe torrado 500g", "qtd": "80", "un": "CX", "valor": "4200.00"},
                ],
            },
            {
                "id": "CTE-2026-0102",
                "tipo": "CTe",
                "numero": "8851",
                "serie": "002",
                "emissao": "2026-02-09 16:08",
                "emitente": "Logistica Norte",
                "destinatario": "Farmacia Vida",
                "valor": "19800.00",
                "status": "EM_ANALISE",
                "documentos": [
                    {
                        "tipo": "MIGO",
                        "numero": "4900007712",
                        "data": "2026-02-09",
                        "status": "CRIADO",
                        "itens": [
                            {"seq": 10, "nfe_item_seq": 1, "material": "FRETE", "descricao": "Servico de transporte", "qtd": "1", "un": "SV", "valor": "19800.00"},
                        ],
                    }
                ],
                "itens": [
                    {"seq": 1, "codigo": "FRETE", "descricao": "Servico de transporte", "qtd": "1", "un": "SV", "valor": "19800.00"},
                ],
            },
            {
                "id": "NFSE-2026-0431",
                "tipo": "NFSe",
                "numero": "431",
                "serie": "A1",
                "emissao": "2026-02-08 11:45",
                "emitente": "Tech Servicios",
                "destinatario": "Industria Nova",
                "valor": "4520.50",
                "status": "PENDENTE",
                "documentos": [],
                "itens": [
                    {"seq": 1, "codigo": "SVC-100", "descricao": "Suporte mensal", "qtd": "1", "un": "SV", "valor": "4520.50"},
                ],
            },
        ]
    }

    return render(request, "manifesto/index.html", {"manifesto_data": manifesto_data})

#--------------------------------------------------------------------
#       Empresas - Modais
#--------------------------------------------------------------------
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
@require_http_methods(["GET","POST"])
def fn_view_inserir_empresa(request):
    """Inserir nova empresa. Superuser pode informar o cliente no formulário."""
    is_superuser = request.session.get('is_superuser', False)
    cod_cliente = request.session.get('cod_cliente', None)
    if request.method == "GET":
        if is_superuser and request.GET.get('cod_cliente'):
            cod_cliente = request.GET.get('cod_cliente', '').strip() or cod_cliente
        if not cod_cliente:
            return JsonResponse({"erro": "Cliente não identificado"}, status=403)
        cl_gdf = ClGdf()
        data = cl_gdf.get_empresa_dados_ins(i_v_cod_cliente=cod_cliente)
        return JsonResponse(data)  
    
    elif request.method == "POST":
        if is_superuser and request.POST.get("m_cod_cliente"):
            cod_cliente = request.POST.get("m_cod_cliente", "").strip() or cod_cliente
        if not cod_cliente:
            return JsonResponse({"erro": "Cliente não identificado"}, status=403)
        cl_gdf = ClGdf()
        cod_empresa = request.POST.get("m_codempresa", "").strip()
        razao = request.POST.get("m_razao", "").strip()
        cnpj = request.POST.get("m_cnpj", "").strip()
        cnpj = re.sub(r"\D", "", cnpj)
        fantasia = request.POST.get("m_fantasia", "").strip()
        matriz = request.POST.get("m_matriz") == "on"
        ie = request.POST.get("m_ie", "").strip()
        im = request.POST.get("m_im", "").strip()
        iest = request.POST.get("m_iest", "").strip()
        crt = request.POST.get("m_crt", "").strip()
        cnae = request.POST.get("m_cnae", "").strip()
        suframa = request.POST.get("m_suframa", "").strip()
        chave_acesso = request.POST.get("m_chave_acesso", "").strip()

        errors = []
        if not cod_empresa:
            errors.append("Código da empresa é obrigatório")
        if not razao:
            errors.append("Razão social é obrigatória")
        if not cnpj:
            errors.append("CNPJ é obrigatório")
        if not fantasia:
            errors.append("Fantasia é obrigatória")
        if errors:
            return JsonResponse({"erro": " | ".join(errors)}, status=400)

        resultado = cl_gdf.set_empresa(
            i_v_cod_empresa=cod_empresa,
            i_v_razao=razao,
            i_v_cnpj=cnpj,
            i_v_fantasia=fantasia,
            i_v_cod_cliente=cod_cliente,
            i_b_matriz=matriz,
            i_v_ie=ie,
            i_v_im=im,
            i_v_iest=iest,
            i_v_crt=crt,
            i_v_cnae=cnae,
            i_v_suframa=suframa,
            i_v_chave_acesso=chave_acesso
        )
        
        # ✅ Verificar resultado
        if not resultado.get("success"):
            return JsonResponse({"erro": resultado.get("message")}, status=400)

        # Garantir que a lista exiba o cliente em que a empresa foi criada (ex.: superuser escolheu outro cliente)
        request.session['cod_cliente'] = cod_cliente
        request.session.modified = True

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": resultado.get("message", "Empresa cadastrada com sucesso")})
        return redirect('Dm_Empresas')


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
@validate_idor_empresa
@require_http_methods(["GET", "POST"])
def fn_view_atualizar_empresa(request, cod_empresa):
    """Atualizar empresa existente"""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    
    cl_gdf = ClGdf()
    if request.method == "GET":
        # Retornar dados da empresa para popular o modal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            empresa_data = cl_gdf.get_empresa_upd(
                i_v_cod_empresa=cod_empresa,
                i_v_cod_cliente=cod_cliente
            )
            return JsonResponse(empresa_data)
        else:
            return JsonResponse({"erro": "Requisição inválida"}, status=400)
    
    elif request.method == "POST":  
        # Atualizar dados da empresa
        razao = request.POST.get("m_razao", "").strip()
        fantasia = request.POST.get("m_fantasia", "").strip()
        ie = request.POST.get("m_ie", "").strip()
        im = request.POST.get("m_im", "").strip()
        iest = request.POST.get("m_iest", "").strip()
        crt = request.POST.get("m_crt", "").strip()
        cnae = request.POST.get("m_cnae", "").strip()
        suframa = request.POST.get("m_suframa", "").strip()
        chave_acesso = request.POST.get("m_chave_acesso", "").strip()
        matriz = request.POST.get("m_matriz") == "on"
        
        resultado = cl_gdf.upd_empresa(
            i_v_cod_empresa=cod_empresa,
            i_v_razao=razao,
            i_v_fantasia=fantasia,
            i_v_ie=ie,
            i_v_im=im,
            i_v_iest=iest,
            i_v_crt=crt,
            i_v_cnae=cnae,
            i_v_suframa=suframa,
            i_v_chave_acesso=chave_acesso,
            i_b_matriz=matriz,
            i_v_cod_cliente=cod_cliente
        )
        
        # ✅ Detectar se é requisição AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if not resultado.get("success"):
            mensagem = resultado.get("message", "Erro ao atualizar")
            if is_ajax:
                return JsonResponse({"success": False, "message": mensagem}, status=400)
            else:
                messages.error(request, mensagem, extra_tags='MODAL_UPD')
                return redirect('Dm_Empresas')
        else:
            mensagem = resultado.get("message", "Empresa atualizada com sucesso!")
            if is_ajax:
                return JsonResponse({"success": True, "message": mensagem}, status=200)
            else:
                messages.success(request, mensagem, extra_tags='MODAL_UPD')
                return redirect('Dm_Empresas')
    
    return JsonResponse({"erro": "Método não permitido"}, status=405)

@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Empresas')
@require_http_methods(["POST"])
def fn_view_atualizar_certificado(request):
    """Atualizar certificado digital da empresa"""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        messages.error(request, "Cliente não identificado.", extra_tags='MODAL_UPD')
        return redirect('Dm_Empresas')

    cod_empresa = request.POST.get('m_codempresa', '').strip()
    if not cod_empresa:
        messages.error(request, "Empresa não identificada.", extra_tags='MODAL_UPD')
        return redirect('Dm_Empresas')

    # ✅ IDOR: empresa deve pertencer ao cliente da sessão
    if not Empresa.objects.filter(cod_empresa=cod_empresa, gdfcliente__cod_cliente=cod_cliente).exists():
        messages.error(request, "Acesso negado: empresa não pertence ao seu cliente.", extra_tags='MODAL_UPD')
        return redirect('Dm_Empresas')

    cl_gdf = ClGdf()

    # ✅ Pegar arquivo (OPCIONAL - pode atualizar só metadados)
    cert_file = request.FILES.get('m_file')
    
    # ✅ Validar extensão só se arquivo foi enviado
    if cert_file and not cert_file.name.endswith(('.crt', '.txt')):
        messages.error(request, "Formato inválido. Use .crt ou .txt", extra_tags='MODAL_UPD')
        return redirect('Dm_Empresas')
    
    # ✅ Extrair dados adicionais do certificado
    emissor = request.POST.get('m_emissor', '').strip()
    cnpj = request.POST.get('m_cnpj', '').strip()
    dt_inicial = request.POST.get('m_dt_inicial', '').strip()
    dt_fim = request.POST.get('m_dt_fim', '').strip()
    senha_certificado = request.POST.get('m_senha_certificado', '').strip()
    
    # ✅ Se nenhum dado foi enviado, erro
    if not cert_file and not (emissor or cnpj or dt_inicial or dt_fim):
        messages.error(request, "Selecione um arquivo ou preencha os dados do certificado", extra_tags='MODAL_UPD')
        return redirect('Dm_Empresas')
    
    # Chamar método de atualização de certificado
    resultado = cl_gdf.upd_certificado(
        i_v_arquivo_cert=cert_file,
        i_v_emissor=emissor,
        i_v_cpf_cnpj=cnpj,
        i_v_ini_validade=dt_inicial,
        i_v_fim_validade=dt_fim,
        i_v_cod_empresa=cod_empresa,
        i_v_senha_certificado=senha_certificado or None,
    )
    
    if resultado.get("success"):
        messages.success(request, resultado.get("message", "Certificado atualizado!"), extra_tags='MODAL_UPD')
    else:
        messages.error(request, resultado.get("message", "Erro ao atualizar certificado"), extra_tags='MODAL_UPD')
    
    return redirect('Dm_Empresas')
#--------------------------------------------------------------------
#       Clientes - Modais
#--------------------------------------------------------------------
@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
@require_http_methods(["GET", "POST"])
def fn_view_inserir_cliente(request):
    """Inserir novo cliente - seguindo padrão Usuario_ins"""
    cl_gdf = ClGdf()
    if request.method == "GET":
        # Retornar dados para preencher o modal (opcionalmente listas de soluções, etc)
        return JsonResponse({"success": True})

    elif request.method == "POST":
        # ✅ Extrair dados do formulário
        cliente_id = request.POST.get("m_cliente_id", "").strip()
        razao = request.POST.get("m_razao", "").strip()
        cnpj = request.POST.get("m_cnpj", "").strip()
        cnpj = re.sub(r"\D", "", cnpj)

        # ✅ Validações básicas
        if not cliente_id:
            return JsonResponse({"erro": "Código do cliente é obrigatório"}, status=400)
        if not razao:
            return JsonResponse({"erro": "Razão social é obrigatória"}, status=400)
        if not cnpj:
            return JsonResponse({"erro": "CNPJ é obrigatório"}, status=400)
        
        # ✅ Chamar método de inserção na classe
        resultado = cl_gdf.set_cliente(
            i_cliente=cliente_id,
            i_razao=razao,
            i_cnpj=cnpj
        )
        
        # ✅ Verificar resultado
        if not resultado.get("success"):
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"erro": resultado.get("message", "Erro ao criar cliente")}, status=400)
            messages.error(request, resultado.get("message", "Erro ao criar cliente"), extra_tags='MODAL_INS')
            return redirect('Dm_Clientes')
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "message": resultado.get("message", "Cliente cadastrado!")})
        messages.success(request, resultado.get("message", "Cliente cadastrado!"), extra_tags='MODAL_INS')
        return redirect('Dm_Clientes')
    
    return JsonResponse({"erro": "Método não permitido"}, status=405)

@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
@require_http_methods(["GET", "POST"])
def fn_view_atualizar_cliente(request, cod_cliente):
    """Atualizar cliente existente - seguindo padrão Usuario_upd"""
    cod_cliente_sessao = request.session.get('cod_cliente', None)
    if not cod_cliente_sessao and not usuario_acesso_total_painel(request):
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    # Usuário normal: só pode editar o cliente da sessão. Acesso total: qualquer cliente.
    if not usuario_acesso_total_painel(request) and str(cod_cliente) != str(cod_cliente_sessao):
        return JsonResponse({"erro": "Acesso negado: você não pode editar outro cliente"}, status=403)
    
    cl_gdf = ClGdf()
    if request.method == "GET":
        # Retornar dados do cliente em JSON (para modal)
        data = cl_gdf.get_cliente_upd(i_v_cliente_id=cod_cliente)
        if not data or data.get('erro'):
            return JsonResponse({"erro": "Cliente não encontrado"}, status=404)

        return JsonResponse(data)

    elif request.method == "POST":
        cod_cliente_id = request.POST.get("upd_cliente_id", "").strip()
        
        # ✅ Extrair dados do formulário
        razao = request.POST.get("upd_razao", "").strip()
        cnpj = request.POST.get("upd_cnpj", "").strip()
        cnpj = re.sub(r"\D", "", cnpj)
        is_active = request.POST.get("upd_is_active") == "on"
        
        # ✅ Validações básicas
        if not razao:
            return JsonResponse({"erro": "Razão social é obrigatória"}, status=400)
        if not cnpj:
            return JsonResponse({"erro": "CNPJ é obrigatório"}, status=400)

        resultado = cl_gdf.upd_cliente(
            i_cliente=cod_cliente_id,
            i_razao=razao,
            i_cnpj=cnpj,
            i_is_active=is_active
        )
        
        # ✅ Detectar se é requisição AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if not resultado.get("success"):
            mensagem = resultado.get("message", "Erro ao atualizar")
            if is_ajax:
                return JsonResponse({"success": False, "message": mensagem}, status=400)
            messages.error(request, mensagem, extra_tags='MODAL_UPD')
            return redirect('Dm_Clientes')

        mensagem = resultado.get("message", "Cliente atualizado com sucesso!")
        if is_ajax:
            return JsonResponse({"success": True, "message": mensagem}, status=200)
        messages.success(request, mensagem, extra_tags='MODAL_UPD')
        return redirect('Dm_Clientes')

@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
@require_http_methods(["POST"])
def fn_view_atualizar_acesso_cliente(request):
    """Atualizar acessos do cliente existente"""
    cod_cliente_sessao = request.session.get('cod_cliente', None)
    if not cod_cliente_sessao and not usuario_acesso_total_painel(request):
        return JsonResponse({"erro": "Cliente não identificado na sessão"}, status=403)
    
    # ✅ Obter o cod_cliente do formulário (cliente sendo editado)
    cod_cliente = request.POST.get("Acesso_cliente_id", "").strip()
    if not cod_cliente:
        messages.error(request, "Cliente não identificado no formulário", extra_tags='MODAL_UPD')
        return redirect('Dm_Clientes')

    # ✅ IDOR: usuário sem acesso total só pode alterar o cliente da sessão
    if not usuario_acesso_total_painel(request) and str(cod_cliente) != str(cod_cliente_sessao):
        messages.error(request, "Acesso negado: você não pode alterar acessos de outro cliente.", extra_tags='MODAL_UPD')
        return redirect('Dm_Clientes')

    cl_gdf = ClGdf()
    ls_solucoes = request.POST.get("ls_solucoes", "").strip()  # Formato: "COD1:1,COD2:0"
    
    print(f"[Cliente_acesso_upd] cod_cliente: {cod_cliente}")
    print(f"[Cliente_acesso_upd] ls_solucoes: {ls_solucoes}")
    
    resultado = cl_gdf.set_cliente_solucoes(
        i_v_cod_cliente=cod_cliente,
        ls_solucoes=ls_solucoes
    )

    print(f"[Cliente_acesso_upd] resultado: {resultado}")

    # ✅ Detectar se é requisição AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not resultado.get("success"):
        mensagem = resultado.get("message", "Erro ao atualizar acessos")
        if is_ajax:
            return JsonResponse({"success": False, "message": mensagem}, status=400)
        else:
            messages.error(request, mensagem, extra_tags='MODAL_UPD')
    else:
        mensagem = resultado.get("message", "Acessos atualizados com sucesso!")
        if is_ajax:
            return JsonResponse({"success": True, "message": mensagem}, status=200)
        else:
            messages.success(request, mensagem, extra_tags='MODAL_UPD')

    if not is_ajax:
        return redirect('Dm_Clientes')


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
@require_http_methods(["POST"])
def fn_view_atualizar_grupos_cliente(request):
    """Atualiza grupos de usuários vinculados ao cliente."""
    cod_cliente_sessao = request.session.get('cod_cliente', None)
    if not cod_cliente_sessao and not usuario_acesso_total_painel(request):
        return JsonResponse({"erro": "Cliente não identificado na sessão"}, status=403)

    cod_cliente = request.POST.get("Grupos_cliente_id", "").strip()
    if not cod_cliente:
        return JsonResponse({"erro": "Cliente não identificado no formulário"}, status=400)

    # ✅ IDOR: usuário sem acesso total só pode alterar o cliente da sessão
    if not usuario_acesso_total_painel(request) and str(cod_cliente) != str(cod_cliente_sessao):
        return JsonResponse({"erro": "Acesso negado: você não pode alterar grupos de outro cliente."}, status=403)

    ls_grupos = request.POST.get("ls_grupos", "").strip()
    cl_gdf = ClGdf()
    resultado = cl_gdf.set_cliente_grupos(
        i_v_cod_cliente=cod_cliente,
        ls_grupos_ids=ls_grupos
    )

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if not resultado.get("success"):
        return JsonResponse({"success": False, "message": resultado.get("message", "Erro ao atualizar grupos")}, status=400)
    return JsonResponse({"success": True, "message": resultado.get("message", "Grupos atualizados com sucesso")}, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes')
@require_http_methods(["POST"])
def fn_view_cliente_sap(request, cod_cliente):
    """Cria ou atualiza a conexão SAP do cliente (uma por cliente)."""
    cod_cliente_sessao = request.session.get('cod_cliente', None)
    if not cod_cliente_sessao and not usuario_acesso_total_painel(request):
        return JsonResponse({"erro": "Cliente não identificado"}, status=403)
    if not usuario_acesso_total_painel(request) and str(cod_cliente) != str(cod_cliente_sessao):
        return JsonResponse({"erro": "Acesso negado."}, status=403)

    try:
        cliente = ClienteGdf.objects.get(cod_cliente=cod_cliente)
    except ClienteGdf.DoesNotExist:
        return JsonResponse({"erro": "Cliente não encontrado"}, status=404)

    sap = ConexaoSap.objects.filter(gdfcliente=cliente).first()
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == "POST":
        ashost = (request.POST.get("sap_ashost") or "").strip()
        sysnr = (request.POST.get("sap_sysnr") or "").strip()
        client = (request.POST.get("sap_client") or "").strip()
        username = (request.POST.get("sap_username") or "").strip()
        passwd = (request.POST.get("sap_passwd") or "").strip()
        lang = (request.POST.get("sap_lang") or "").strip()
        active = request.POST.get("sap_active") == "on"

        if not sap:
            sap = ConexaoSap.objects.create(
                gdfcliente=cliente,
                ashost=ashost,
                sysnr=sysnr,
                client=client,
                username=username,
                passwd=passwd,
                lang=lang,
                active=active,
            )
            message = "Conexão SAP criada. Preencha os dados e salve novamente."
        else:
            sap.ashost = ashost
            sap.sysnr = sysnr
            sap.client = client
            sap.username = username
            if passwd:
                sap.passwd = passwd
            sap.lang = lang
            sap.active = active
            sap.save()
            message = "Conexão SAP atualizada com sucesso."

        sap_data = {
            "id": sap.id,
            "ashost": sap.ashost or "",
            "sysnr": sap.sysnr or "",
            "client": sap.client or "",
            "username": sap.username or "",
            "passwd": sap.passwd or "",
            "lang": sap.lang or "",
            "active": sap.active,
        }
        if is_ajax:
            return JsonResponse({
                "success": True,
                "message": message,
                "sap_connection": sap_data,
            }, status=200)
        messages.success(request, message, extra_tags='MODAL_UPD')
        return redirect('Dm_Clientes')
    return JsonResponse({"erro": "Método não permitido"}, status=405)


def _contexto_painel_relatorio_fiscal(request):
    """
    Dados de filtros e opções do painel de relatório (NFe, CTe, NFS, SPED).
    Usado nas páginas Carga XML e Carga SPED; sem cliente na sessão retorna dict mínimo.
    """
    cod_cliente = request.session.get("cod_cliente", None)
    if not cod_cliente:
        return {
            "empresas_usuario": [],
            "filiais_usuario": Filial.objects.none(),
            "meio_pagamento_choices": [],
            "relatorio_condicao_sap_opcoes": [],
        }
    try:
        cliente = ClienteGdf.objects.get(cod_cliente=cod_cliente)
        if usuario_acesso_total_painel(request):
            empresas_usuario = Empresa.objects.filter(gdfcliente=cliente).order_by("fantasia", "razao", "cod_empresa").distinct()
        else:
            empresas_usuario = Empresa.objects.filter(
                gdfcliente=cliente, usuarioempresa__user=request.user
            ).order_by("fantasia", "razao", "cod_empresa").distinct()
    except ClienteGdf.DoesNotExist:
        return {
            "empresas_usuario": [],
            "filiais_usuario": Filial.objects.none(),
            "meio_pagamento_choices": [],
            "relatorio_condicao_sap_opcoes": [],
        }
    relatorio_condicao_sap_opcoes = list(
        CondicaoParam.objects.filter(gdfcliente_id=cod_cliente)
        .exclude(condicao_pagamento_sap__isnull=True)
        .exclude(condicao_pagamento_sap="")
        .values_list("condicao_pagamento_sap", flat=True)
        .distinct()
        .order_by("condicao_pagamento_sap")[:200]
    )
    try:
        meio_pagamento_choices = list(NFe_Pagamento._meta.get_field("meio_pagamento").choices)
    except Exception:
        meio_pagamento_choices = []
    filiais_usuario = (
        Filial.objects.filter(empresa__in=empresas_usuario)
        .select_related("empresa")
        .order_by("empresa__fantasia", "empresa__razao", "empresa__cod_empresa", "cod_filial")
    )
    return {
        "empresas_usuario": empresas_usuario,
        "filiais_usuario": filiais_usuario,
        "meio_pagamento_choices": meio_pagamento_choices,
        "relatorio_condicao_sap_opcoes": relatorio_condicao_sap_opcoes,
    }


def _usuario_pode_relatorio_excel_nfe_cte_nfse(request):
    subs = get_subsolucoes_usuario(request.user)
    if subs is None:
        return True
    return "Pro_CargaXml" in subs


def _usuario_pode_relatorio_excel_sped(request):
    subs = get_subsolucoes_usuario(request.user)
    if subs is None:
        return True
    return "Pro_CargaSped" in subs


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml')
@require_http_methods(["GET"])
def fn_view_CargaXml(request):
    """View para carregamento de XML"""
    cod_cliente = request.session.get('cod_cliente', None)
    
    if not cod_cliente:
        return render(request, 'comum/login.html', {'error_message': 'Cliente não identificado'})
    
    # Buscar jobs do cliente (todos os registros)
    try:
        cliente = ClienteGdf.objects.get(cod_cliente=cod_cliente)
        jobs = (
            JobCargaXml.objects.filter(gdfcliente=cliente)
            .select_related("gdfcliente")
            .order_by("-started_at")
        )
    except ClienteGdf.DoesNotExist:
        jobs = []
    url_prefix = (request.META.get("SCRIPT_NAME") or getattr(settings, "FORCE_SCRIPT_NAME", "") or "").strip()
    if url_prefix and not url_prefix.startswith("/"):
        url_prefix = "/" + url_prefix
    url_prefix = url_prefix.rstrip("/")  # '' ou '/gdf'

    context = {
        "cod_cliente": cod_cliente,
        "jobs": jobs,
        "url_prefix": url_prefix,
        "tipo_pagamento_desc": TIPO_PAGAMENTO_DESC,
        "relatorio_painel": "xml",
    }
    context.update(_contexto_painel_relatorio_fiscal(request))
    return render(request, "importacao/index_carga_xml.html", context)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_processar_xml(request):
    """API para processar upload de XMLs em segundo plano (job)."""
    cod_cliente = request.session.get('cod_cliente', None)
    
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    
    try:
        lsl_Xml = request.FILES.getlist('arquivo')
        l_v_type_xml = (request.POST.get('type_xml') or 'NFe').strip() or 'NFe'
        l_v_empresa_id = (request.POST.get('empresa_id') or '').strip()

        if not lsl_Xml:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhum arquivo selecionado. Envie arquivos .xml ou .zip (campo "arquivo"). Se escolheu uma pasta, confira se há arquivos .xml dentro.'
            }, status=400)

        # Aceitar .xml e .zip; se for .zip, extrair XMLs na "mesma pasta" (lista) e processar
        MAX_SIZE = 50 * 1024 * 1024
        expanded = []
        for f in lsl_Xml:
            name_lower = f.name.lower()
            if name_lower.endswith('.xml'):
                if f.size > MAX_SIZE:
                    return JsonResponse({'sucesso': False, 'mensagem': f'Arquivo muito grande: {f.name}'}, status=400)
                expanded.append(f)
            elif name_lower.endswith('.zip'):
                if f.size > MAX_SIZE:
                    return JsonResponse({'sucesso': False, 'mensagem': f'Arquivo ZIP muito grande: {f.name}'}, status=400)
                try:
                    if hasattr(f, 'seek'):
                        f.seek(0)
                    # ZIP: suporte a nomes no encoding do Windows (CP437) quando disponível (Python 3.11+)
                    try:
                        zf = zipfile.ZipFile(f, 'r', metadata_encoding='cp437')
                    except TypeError:
                        zf = zipfile.ZipFile(f, 'r')
                    with zf:
                        seen_basenames = {}
                        for member in zf.namelist():
                            # Normalizar caminho (ZIPs do Windows usam \)
                            member_norm = member.replace('\\', '/').strip()
                            # Só processar .xml; ignorar pastas, .html, .pdf, .txt, etc.
                            if member_norm.endswith('/') or not member_norm.lower().endswith('.xml'):
                                continue
                            safe_name = os.path.basename(member_norm) or os.path.basename(member.replace('\\', '/')) or 'arquivo.xml'
                            # Nomes únicos quando o ZIP tem pastas com arquivos de mesmo nome
                            if safe_name in seen_basenames:
                                seen_basenames[safe_name] += 1
                                base, ext = os.path.splitext(safe_name)
                                safe_name = f"{base}_{seen_basenames[safe_name]}{ext}"
                            else:
                                seen_basenames[safe_name] = 1
                            with zf.open(member, 'r') as src:
                                xml_bytes = src.read()
                            if len(xml_bytes) > MAX_SIZE:
                                return JsonResponse({'sucesso': False, 'mensagem': f'XML dentro do ZIP muito grande: {safe_name}'}, status=400)
                            expanded.append(SimpleUploadedFile(safe_name, xml_bytes))
                except (zipfile.BadZipFile, OSError, ValueError) as e:
                    return JsonResponse({'sucesso': False, 'mensagem': f'ZIP inválido ou corrompido: {f.name} - {e}'}, status=400)
            else:
                continue  # Ignorar arquivos que não sejam .xml ou .zip (ex.: PDF em pasta)

        if not expanded:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhum arquivo XML encontrado. Envie apenas arquivos .xml ou .zip que contenham .xml dentro (outros tipos são ignorados).'
            }, status=400)

        max_arquivos_por_requisicao = getattr(settings, 'CARGAXML_MAX_ARCHIVOS_POR_REQUISICAO', 5000)
        if len(expanded) > max_arquivos_por_requisicao:
            return JsonResponse({
                'sucesso': False,
                'mensagem': f'Máximo de {max_arquivos_por_requisicao} arquivos por envio. Você enviou {len(expanded)}. Compacte em .zip ou envie em partes.'
            }, status=400)

        if l_v_type_xml not in ('NFe', 'CTe', 'NFSe'):
            return JsonResponse({
                'sucesso': False,
                'mensagem': f'Tipo de documento inválido: {l_v_type_xml}. Use NFe, CTe ou NFSe.'
            }, status=400)

        def _temp_dir_para_job(jid):
            base = getattr(settings, 'CARGAXML_TEMP_ROOT', None) or tempfile.gettempdir()
            return os.path.join(base, 'gdf_cargaxml', str(jid))

        cliente = get_object_or_404(ClienteGdf, cod_cliente=cod_cliente)
        from app.api.jobs import processar_job_xml_background

        def _disparar_processamento_xml(job, temp_dir):
            """Dispara processamento em Celery ou, se indisponível, em thread."""
            try:
                from app.api.tasks import processar_job_xml_manual
                processar_job_xml_manual.delay(
                    job.id, temp_dir, l_v_type_xml,
                    request.user.id, cod_cliente, l_v_empresa_id or None,
                )
                return
            except Exception:
                pass
            t = threading.Thread(
                target=processar_job_xml_background,
                args=(job.id, temp_dir, l_v_type_xml, request.user.id, cod_cliente, l_v_empresa_id),
                daemon=True,
            )
            t.start()

        # Um envio = um job: criar job, salvar arquivos na pasta temp e disparar processamento
        job = JobCargaXml.objects.create(
            gdfcliente=cliente,
            status='RUNNING',
            total_arquivos=len(expanded),
            total_sucesso=0,
            total_erro=0,
            mensagem=f'Carga manual – em execução ({len(expanded)} arquivo(s))...',
            started_at=timezone.localtime(),
            finished_at=None,
            usuario_execucao=request.user,
        )
        temp_dir = _temp_dir_para_job(job.id)
        try:
            os.makedirs(temp_dir, exist_ok=True)
            for i, item in enumerate(expanded):
                xml_bytes = item.read() if hasattr(item, 'read') else item
                nome = getattr(item, 'name', f'{i}.xml')
                safe_name = os.path.basename(nome).replace('..', '_') or f'{i}.xml'
                dest = os.path.join(temp_dir, f'{i}_{safe_name}')
                with open(dest, 'wb') as out:
                    out.write(xml_bytes)
        except Exception as e:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception:
                pass
            job.delete()
            return JsonResponse({'sucesso': False, 'mensagem': f'Erro ao salvar arquivos: {e}'}, status=500)

        _disparar_processamento_xml(job, temp_dir)
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Job #{job.id} criado e em execução em segundo plano ({len(expanded)} arquivo(s)). Atualize o painel para acompanhar.',
            'job_id': job.id,
        }, status=202)

    except Exception as e:
        return JsonResponse({'sucesso': False, 'mensagem': f'Erro ao processar: {str(e)}'}, status=500)



@login_required(login_url='Login')
@requer_acesso_total_painel(redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_sessao_cliente(request):
    """Define o cliente ativo na sessão. Apenas superuser ou cliente PRCIT (dona do projeto)."""
    try:
        # Aceita FormData (request.POST) ou JSON (request.body) - ler apenas um para evitar RawPostDataException
        ct = (request.content_type or '').lower()
        if 'application/json' in ct:
            body = json.loads(request.body) if request.body else {}
            cod_cliente = (body.get('cod_cliente') or '').strip()
        else:
            cod_cliente = (request.POST.get('cod_cliente') or '').strip()
        if not cod_cliente:
            return JsonResponse({'sucesso': False, 'erro': 'cod_cliente obrigatório'}, status=400)
        if not ClienteGdf.objects.filter(cod_cliente=cod_cliente, is_active=True).exists():
            return JsonResponse({'sucesso': False, 'erro': 'Cliente não encontrado ou inativo'}, status=400)
        request.session['cod_cliente'] = cod_cliente
        return JsonResponse({'sucesso': True, 'cod_cliente': cod_cliente}, status=200)
    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido'}, status=400)


@login_required(login_url='Login')
@requer_superuser(redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_debug_session(request):
    """Debug endpoint para verificar sessão e cliente (apenas superuser)."""
    cod_cliente = request.session.get('cod_cliente', None)
    return JsonResponse({
        'usuario': request.user.username,
        'cod_cliente': cod_cliente,
        'session_keys': list(request.session.keys()),
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargaxml_avisos(request):
    """Retorna jobs de carga XML com status ERROR (para o botão Avisos e modal de logs)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente nao identificado'}, status=403)
    jobs = JobCargaXml.objects.filter(
        gdfcliente__cod_cliente=cod_cliente,
        status='ERROR'
    ).order_by('-started_at')[:100]
    items = []
    def _ordem_log_erros_primeiro(lines):
        def prioridade(line):
            t = (line or '').strip()
            if t.startswith('ERRO:'):
                return 0
            if t.startswith('PENDENTES'):
                return 1
            if t.startswith('OK:'):
                return 2
            return 3
        return sorted(lines, key=prioridade)

    for job in jobs:
        log_lines = [line.strip() for line in (job.mensagem or '').splitlines() if line.strip()]
        log_lines = _ordem_log_erros_primeiro(log_lines)
        items.append({
            'id': job.id,
            'started_at': isoformat_brasilia(job.started_at),
            'finished_at': isoformat_brasilia(job.finished_at),
            'total_arquivos': job.total_arquivos,
            'total_sucesso': job.total_sucesso,
            'total_erro': job.total_erro,
            'mensagem': job.mensagem or '',
            'log': log_lines,
        })
    return JsonResponse({'sucesso': True, 'total_erros': len(items), 'items': items}, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargaxml_jobs(request):
    """Lista todos os jobs de carga XML do cliente (inclui mensagem para monitoramento)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente nao identificado'}, status=403)

    jobs = JobCargaXml.objects.filter(gdfcliente__cod_cliente=cod_cliente).order_by('-started_at')
    items = []
    for job in jobs:
        items.append({
            'id': job.id,
            'status': job.status,
            'total_arquivos': job.total_arquivos,
            'total_sucesso': job.total_sucesso,
            'total_erro': job.total_erro,
            'mensagem': (job.mensagem or '')[:500],
            'started_at': isoformat_brasilia(job.started_at),
            'finished_at': isoformat_brasilia(job.finished_at),
        })
    return JsonResponse({'sucesso': True, 'items': items}, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargaxml_resumo(request):
    """Retorna contagens para o painel: total de jobs, concluídos, com erros, em andamento."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    qs = JobCargaXml.objects.filter(gdfcliente__cod_cliente=cod_cliente)
    total = qs.count()
    concluidos = qs.filter(status='SUCCESS').count()
    com_erros = qs.filter(status='ERROR').count()
    em_andamento = qs.filter(status__in=('RUNNING', 'PENDING')).count()
    return JsonResponse({
        'sucesso': True,
        'total': total,
        'concluidos': concluidos,
        'com_erros': com_erros,
        'em_andamento': em_andamento,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargaxml_job_details(request, job_id):
    """Retorna detalhes e log de um job específico"""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente nao identificado'}, status=403)

    job = get_object_or_404(JobCargaXml, id=job_id, gdfcliente__cod_cliente=cod_cliente)
    log_lines = [line.strip() for line in (job.mensagem or '').splitlines() if line.strip()]

    def _prioridade_log(line):
        t = (line or '').strip()
        if t.startswith('ERRO:'):
            return 0
        if t.startswith('PENDENTES'):
            return 1
        if t.startswith('OK:'):
            return 2
        return 3
    log_lines = sorted(log_lines, key=_prioridade_log)

    return JsonResponse({
        'sucesso': True,
        'job': {
            'id': job.id,
            'status': job.status,
            'total_arquivos': job.total_arquivos,
            'total_sucesso': job.total_sucesso,
            'total_erro': job.total_erro,
            'started_at': isoformat_brasilia(job.started_at),
            'finished_at': isoformat_brasilia(job.finished_at),
        },
        'parametro': None,
        'log': log_lines,
    }, status=200)


# ========== APIs Carga SPED (mesma linha de raciocínio da Carga XML) ==========


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_processar_sped(request):
    """API para processar upload de arquivos SPED (.txt). Suporta lotes em um único job (job_id + ultimo_lote)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    arquivos_raw = request.FILES.getlist('arquivo')
    if not arquivos_raw and request.FILES.get('arquivo'):
        arquivos_raw = [request.FILES.get('arquivo')]
    if not arquivos_raw:
        return JsonResponse({'sucesso': False, 'mensagem': 'Nenhum arquivo selecionado. Selecione arquivos .txt ou uma pasta.'}, status=400)

    arquivos = [f for f in arquivos_raw if getattr(f, 'name', None) and (f.name or '').lower().endswith('.txt')]
    if not arquivos:
        return JsonResponse({'sucesso': False, 'mensagem': 'Nenhum arquivo .txt encontrado. Envie apenas arquivos .txt.'}, status=400)

    job_id_existente = request.POST.get('job_id', '').strip()
    ultimo_lote = request.POST.get('ultimo_lote', '').strip().lower() in ('1', 'true', 's', 'sim', 'yes')

    def _temp_dir_para_job_sped(jid):
        base = getattr(settings, 'CARGASPED_TEMP_ROOT', None) or tempfile.gettempdir()
        return os.path.join(base, 'gdf_cargasped', str(jid))

    cliente = get_object_or_404(ClienteGdf, cod_cliente=cod_cliente)
    from app.api.jobs import processar_job_sped_background

    if job_id_existente:
        try:
            job_id_int = int(job_id_existente)
        except ValueError:
            return JsonResponse({'sucesso': False, 'mensagem': 'job_id inválido.'}, status=400)
        job = get_object_or_404(JobCargaSped, id=job_id_int, gdfcliente__cod_cliente=cod_cliente)
        if job.status != 'PENDING':
            return JsonResponse({'sucesso': False, 'mensagem': f'Job #{job.id} já foi finalizado ou está em execução (status={job.status}).'}, status=400)
        temp_dir = _temp_dir_para_job_sped(job.id)
        if not os.path.isdir(temp_dir):
            return JsonResponse({'sucesso': False, 'mensagem': f'Pasta do job #{job.id} não encontrada. Envie os lotes em sequência.'}, status=400)
        offset = len([f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f)) and f.lower().endswith('.txt')])
        try:
            for i, f in enumerate(arquivos):
                safe_name = os.path.basename(f.name).replace('..', '_')
                dest = os.path.join(temp_dir, f'{offset + i}_{safe_name}')
                with open(dest, 'wb') as out:
                    for chunk in f.chunks():
                        out.write(chunk)
        except Exception as e:
            return JsonResponse({'sucesso': False, 'mensagem': f'Erro ao salvar arquivos no job: {e}'}, status=500)
        job.total_arquivos += len(arquivos)
        job.save(update_fields=['total_arquivos'])
        if ultimo_lote:
            job.status = 'RUNNING'
            job.mensagem = f'Carga manual – em execução ({job.total_arquivos} arquivo(s))...'
            job.started_at = timezone.localtime()
            job.save(update_fields=['status', 'mensagem', 'started_at'])
            t = threading.Thread(
                target=processar_job_sped_background,
                args=(job.id, temp_dir, cod_cliente, request.user.id),
                daemon=True,
            )
            t.start()
            return JsonResponse({
                'sucesso': True,
                'mensagem': f'Job #{job.id} finalizado e em execução ({job.total_arquivos} arquivo(s)). Atualize o painel.',
                'job_id': job.id,
            }, status=202)
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Lote recebido. Job #{job.id} com {job.total_arquivos} arquivo(s). Envie o próximo lote com o mesmo job_id.',
            'job_id': job.id,
        }, status=200)

    job = JobCargaSped.objects.create(
        gdfcliente=cliente,
        status='PENDING',
        total_arquivos=len(arquivos),
        total_sucesso=0,
        total_erro=0,
        mensagem='Aguardando lotes...',
        started_at=None,
        finished_at=None,
        usuario_execucao=request.user,
    )
    temp_dir = _temp_dir_para_job_sped(job.id)
    try:
        os.makedirs(temp_dir, exist_ok=True)
        for i, f in enumerate(arquivos):
            safe_name = os.path.basename(f.name).replace('..', '_')
            dest = os.path.join(temp_dir, f'{i}_{safe_name}')
            with open(dest, 'wb') as out:
                for chunk in f.chunks():
                    out.write(chunk)
    except Exception as e:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass
        job.delete()
        return JsonResponse({'sucesso': False, 'mensagem': f'Erro ao salvar arquivos: {e}'}, status=500)

    if ultimo_lote:
        job.status = 'RUNNING'
        job.mensagem = f'Carga manual – em execução ({job.total_arquivos} arquivo(s))...'
        job.started_at = timezone.localtime()
        job.save(update_fields=['status', 'mensagem', 'started_at'])
        t = threading.Thread(
            target=processar_job_sped_background,
            args=(job.id, temp_dir, cod_cliente, request.user.id),
            daemon=True,
        )
        t.start()
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Job #{job.id} criado e em execução. Atualize o painel para acompanhar.',
            'job_id': job.id,
        }, status=202)
    return JsonResponse({
        'sucesso': True,
        'mensagem': f'Job #{job.id} criado com {job.total_arquivos} arquivo(s). Envie o próximo lote com job_id={job.id} e ultimo_lote=1 no último.',
        'job_id': job.id,
    }, status=200)




@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargasped_resumo(request):
    """Retorna contagens para o painel: total de jobs, concluídos, com erros, em andamento."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    qs = JobCargaSped.objects.filter(gdfcliente__cod_cliente=cod_cliente)
    total = qs.count()
    concluidos = qs.filter(status='SUCCESS').count()
    com_erros = qs.filter(status='ERROR').count()
    em_andamento = qs.filter(status__in=('RUNNING', 'PENDING')).count()
    return JsonResponse({
        'sucesso': True,
        'total': total,
        'concluidos': concluidos,
        'com_erros': com_erros,
        'em_andamento': em_andamento,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargasped_avisos(request):
    """Retorna jobs de carga SPED com status ERROR (para o botão Avisos e modal de logs)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    jobs = JobCargaSped.objects.filter(
        gdfcliente__cod_cliente=cod_cliente,
        status='ERROR'
    ).order_by('-started_at')[:100]
    items = []
    for job in jobs:
        msg = job.mensagem or ''
        log_lines = [line.strip() for line in msg.splitlines() if line.strip()]
        if not log_lines and msg.strip():
            log_lines = [msg.strip()]
        items.append({
            'id': job.id,
            'started_at': isoformat_brasilia(job.started_at),
            'finished_at': isoformat_brasilia(job.finished_at),
            'total_arquivos': job.total_arquivos,
            'total_sucesso': job.total_sucesso,
            'total_erro': job.total_erro,
            'mensagem': msg,
            'log': log_lines,
        })
    return JsonResponse({'sucesso': True, 'total_erros': len(items), 'items': items}, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargasped_jobs(request):
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    jobs = JobCargaSped.objects.filter(gdfcliente__cod_cliente=cod_cliente).order_by('-started_at')
    items = [{
        'id': j.id,
        'status': j.status,
        'total_arquivos': j.total_arquivos,
        'total_sucesso': j.total_sucesso,
        'total_erro': j.total_erro,
        'started_at': isoformat_brasilia(j.started_at),
        'finished_at': isoformat_brasilia(j.finished_at),
    } for j in jobs]
    return JsonResponse({'sucesso': True, 'items': items}, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_cargasped_job_details(request, job_id):
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    job = get_object_or_404(JobCargaSped, id=job_id, gdfcliente__cod_cliente=cod_cliente)
    log_lines = [line.strip() for line in (job.mensagem or '').splitlines() if line.strip()]
    return JsonResponse({
        'sucesso': True,
        'job': {
            'id': job.id, 'status': job.status,
            'total_arquivos': job.total_arquivos,
            'total_sucesso': job.total_sucesso,
            'total_erro': job.total_erro,
            'started_at': isoformat_brasilia(job.started_at),
            'finished_at': isoformat_brasilia(job.finished_at),
        },
        'parametro': None,
        'log': log_lines,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped')
@require_http_methods(["GET"])
def fn_view_CargaSped(request):
    """View para carregamento de arquivos SPED (mesma linha de raciocínio da Carga XML)."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return render(request, 'comum/login.html', {'error_message': 'Cliente não identificado'})
    try:
        cliente = ClienteGdf.objects.get(cod_cliente=cod_cliente)
        jobs = (
            JobCargaSped.objects.filter(gdfcliente=cliente)
            .select_related("gdfcliente")
            .order_by("-started_at")
        )
    except ClienteGdf.DoesNotExist:
        jobs = []
    url_prefix = (request.META.get("SCRIPT_NAME") or getattr(settings, "FORCE_SCRIPT_NAME", "") or "").strip()
    if url_prefix and not url_prefix.startswith("/"):
        url_prefix = "/" + url_prefix
    url_prefix = url_prefix.rstrip("/")  # '' ou '/gdf'
    context = {
        "cod_cliente": cod_cliente,
        "jobs": jobs,
        "url_prefix": url_prefix,
        "tipo_pagamento_desc": TIPO_PAGAMENTO_DESC,
        "relatorio_painel": "sped",
    }
    context.update(_contexto_painel_relatorio_fiscal(request))
    return render(request, "importacao/index_carga_sped.html", context)


# ========== APIs Relatório Fiscal (NFe, CTe, NFS, SPED nível cabeçalho) ==========

@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_nfe(request):
    """Lista NFe nível cabeçalho com filtros empresa, grupo de empresa e período."""
    try:
        params = parse_relatorio_params(request, relatorio_empresas_queryset)
    except ValidationError:
        return JsonResponse({'erro': 'Parâmetro de busca inválido'}, status=400)
    if not params.cod_empresas and not params.cod_cliente:
        return JsonResponse({'sucesso': True, 'items': []}, status=200)

    qs = queryset_relatorio_nfe(request, params)
    total, total_pages, page, qs = paginate_queryset(qs, params.page, params.page_size)
    qs = qs.select_related('identificacao__pagamento').prefetch_related(
        Prefetch(
            'identificacao__cobranca__parcelas',
            queryset=NFe_Parcela.objects.order_by('numero_parcela'),
        )
    )
    cod_cli_param = (params.cod_cliente or '').strip() or None
    items = []
    for nfe in qs:
        id_ = nfe.identificacao
        cond_nfe = condicao_pagamento_da_nfe(id_)
        tipo_pag = tipo_pagamento_da_nfe(id_)
        cond_sap = (
            _condicao_sap_da_param(cond_nfe, tipo_pagamento=tipo_pag, cod_cliente=cod_cli_param)
            if cod_cli_param
            else ''
        )
        items.append({
            'id_nfe': nfe.id_nfe,
            'numero': id_.numero,
            'serie': id_.serie,
            'chave': id_.chave_acesso,
            'emissao': id_.emissao.isoformat() if id_.emissao else None,
            'tipo_operacao': id_.tipo_operacao,
            'status': nfe.status,
            'empresa': nfe.empresa.cod_empresa if nfe.empresa else None,
            'natureza': id_.natureza_operacao,
            'filial': nfe.filial.cod_filial if nfe.filial else None,
            'filial_nome': (nfe.filial.nome or '') if nfe.filial else '',
            'condicao_pagamento_sap': cond_sap,
            'tem_sap': nfe.tem_sap,
            'sap_nome_tabela': nfe.sap_nome_tabela or '',
        })
    return JsonResponse({
        'sucesso': True,
        'items': items,
        'total': total,
        'page': page,
        'page_size': params.page_size,
        'total_pages': total_pages,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_cte(request):
    """Lista CTe nível cabeçalho com filtros empresa, grupo de empresa e período."""
    try:
        params = parse_relatorio_params(request, relatorio_empresas_queryset)
    except ValidationError:
        return JsonResponse({'erro': 'Parâmetro de busca inválido'}, status=400)
    if not params.cod_empresas and not params.cod_cliente:
        return JsonResponse({'sucesso': True, 'items': []}, status=200)

    qs = queryset_relatorio_cte(request, params)
    total, total_pages, page, qs = paginate_queryset(qs, params.page, params.page_size)
    items = []
    for cte in qs:
        id_ = cte.identificacao
        items.append({
            'id_cte': cte.id_cte,
            'numero': id_.numero,
            'serie': id_.serie,
            'chave': id_.chave_acesso,
            'emissao': id_.emissao.isoformat() if id_.emissao else None,
            'empresa': cte.empresa.cod_empresa if cte.empresa else None,
            'filial': cte.filial.cod_filial if cte.filial else None,
            'filial_nome': (cte.filial.nome or '') if cte.filial else '',
            'tem_sap': cte.tem_sap,
            'sap_nome_tabela': cte.sap_nome_tabela or '',
        })
    return JsonResponse({
        'sucesso': True,
        'items': items,
        'total': total,
        'page': page,
        'page_size': params.page_size,
        'total_pages': total_pages,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_nfse(request):
    """Lista NFSe nível cabeçalho com filtros empresa, grupo de empresa e período."""
    try:
        params = parse_relatorio_params(request, relatorio_empresas_queryset)
    except ValidationError:
        return JsonResponse({'erro': 'Parâmetro de busca inválido'}, status=400)
    if not params.cod_empresas and not params.cod_cliente:
        return JsonResponse({'sucesso': True, 'items': []}, status=200)

    qs = queryset_relatorio_nfse(request, params)
    total, total_pages, page, qs = paginate_queryset(qs, params.page, params.page_size)
    items = []
    for nfse in qs:
        id_ = nfse.identificacao
        items.append({
            'id_nfse': nfse.id_nfse,
            'numero': id_.numero,
            'chave': id_.chave,
            'emissao': id_.emissao.isoformat() if id_.emissao else None,
            'empresa': nfse.empresa.cod_empresa if nfse.empresa else None,
            'filial': nfse.filial.cod_filial if nfse.filial else None,
            'filial_nome': (nfse.filial.nome or '') if nfse.filial else '',
            'tem_sap': nfse.tem_sap,
            'sap_nome_tabela': nfse.sap_nome_tabela or '',
        })
    return JsonResponse({
        'sucesso': True,
        'items': items,
        'total': total,
        'page': page,
        'page_size': params.page_size,
        'total_pages': total_pages,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_sped(request):
    """Lista SPED nível cabeçalho. tipo_sped: C=Contribuição, F=Fiscal. Busca em sped_fiscal e sped_contribuicao."""
    try:
        params = parse_relatorio_params(request, relatorio_empresas_queryset)
    except ValidationError:
        return JsonResponse({'erro': 'Parâmetro de busca inválido'}, status=400)
    if not params.cod_empresas and not params.cod_cliente:
        return JsonResponse({'sucesso': True, 'items': []}, status=200)

    items = list_relatorio_sped_items(request, params)
    total = len(items)
    page_size = params.page_size
    page = params.page
    total_pages = max(1, (total + page_size - 1) // page_size) if total else 1
    page = min(page, total_pages)
    start = (page - 1) * page_size
    items = items[start : start + page_size]
    return JsonResponse({
        'sucesso': True,
        'items': items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }, status=200)


@login_required(login_url='Login')
@require_http_methods(['GET'])
def fn_api_relatorio_excel(request):
    """
    Exporta planilha .xlsx (aba Resumo + dados) com os mesmos filtros das APIs
    de relatório (sem paginação).

    Query opcional ``excel_tipo``: ``todos`` (padrão, todas as abas),
    ``importacao`` (NFe, CT-e e NFS-e, sem SPED), ``nfe``, ``cte``,
    ``nfse`` ou ``sped`` — define quais abas de dados são geradas.
    """
    try:
        params = parse_relatorio_params(request, relatorio_empresas_queryset)
    except ValidationError:
        return JsonResponse({'erro': 'Parâmetro de busca inválido'}, status=400)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse(
            {
                'sucesso': False,
                'mensagem': 'Exportação Excel indisponível: instale o pacote openpyxl (pip install openpyxl).',
            },
            status=503,
        )

    MAX_ROWS = 100000
    excel_tipo = (request.GET.get('excel_tipo') or 'todos').strip().lower()
    if excel_tipo not in ('todos', 'nfe', 'cte', 'nfse', 'sped', 'importacao'):
        excel_tipo = 'todos'
    incluir_nfe = excel_tipo in ('todos', 'nfe', 'importacao')
    incluir_cte = excel_tipo in ('todos', 'cte', 'importacao')
    incluir_nfse = excel_tipo in ('todos', 'nfse', 'importacao')
    incluir_sped = excel_tipo in ('todos', 'sped')
    if (incluir_nfe or incluir_cte or incluir_nfse) and not _usuario_pode_relatorio_excel_nfe_cte_nfse(request):
        return JsonResponse(
            {"sucesso": False, "mensagem": "Sem permissão para exportar NFe, CT-e ou NFS-e."},
            status=403,
        )
    if incluir_sped and not _usuario_pode_relatorio_excel_sped(request):
        return JsonResponse(
            {"sucesso": False, "mensagem": "Sem permissão para exportar SPED."},
            status=403,
        )

    wb = Workbook()
    bold = Font(bold=True)
    sheets_widths = []

    def _style_header(ws_row1):
        for cell in ws_row1:
            cell.font = bold

    # --- Resumo dos filtros ---
    ws0 = wb.active
    ws0.title = 'Resumo'
    ws0.append(['Parâmetro', 'Valor'])
    _style_header(ws0[1])
    filtros = [
        ('cod_cliente', params.cod_cliente or ''),
        ('empresa_id', params.empresa_id or ''),
        ('data_inicio', params.data_inicio or ''),
        ('data_fim', params.data_fim or ''),
        ('busca', params.busca or ''),
        ('tem_sap', params.tem_sap or ''),
        ('tipo_sped', (request.GET.get('tipo_sped') or '').strip()),
        ('tipo_operacao', (request.GET.get('tipo_operacao') or '').strip()),
        ('tipo_pagamento', (request.GET.get('tipo_pagamento') or '').strip()),
        ('parcelas', (request.GET.get('parcelas') or '').strip()),
        ('condicao_pagamento_sap', (request.GET.get('condicao_pagamento_sap') or '').strip()),
        ('filial_id', (request.GET.get('filial_id') or '').strip()),
        ('order', (request.GET.get('order') or '').strip()),
        ('dir', (request.GET.get('dir') or '').strip()),
        ('excel_tipo', excel_tipo),
        ('limite_linhas_por_aba', str(MAX_ROWS)),
    ]
    for k, v in filtros:
        ws0.append([k, v])
    ws0.column_dimensions['A'].width = 28
    ws0.column_dimensions['B'].width = 72

    cod_cli_param = (params.cod_cliente or '').strip() or None

    if incluir_nfe:
        ws_nfe = wb.create_sheet('NFe')
        h_nfe = [
            'id_nfe',
            'Número',
            'Série',
            'Chave',
            'Emissão',
            'Tipo operação',
            'Status',
            'Empresa',
            'Filial',
            'Nome filial',
            'Natureza',
            'Condição pagamento SAP',
            'Chave no SAP',
            'Tabela SAP',
        ]
        ws_nfe.append(h_nfe)
        _style_header(ws_nfe[1])
        qs_nfe = queryset_relatorio_nfe(request, params).select_related('identificacao__pagamento').prefetch_related(
            Prefetch(
                'identificacao__cobranca__parcelas',
                queryset=NFe_Parcela.objects.order_by('numero_parcela'),
            )
        )
        n_nfe = 0
        for nfe in qs_nfe.iterator(chunk_size=500):
            if n_nfe >= MAX_ROWS:
                break
            id_ = nfe.identificacao
            cond_nfe = condicao_pagamento_da_nfe(id_)
            tipo_pag = tipo_pagamento_da_nfe(id_)
            cond_sap = (
                _condicao_sap_da_param(cond_nfe, tipo_pagamento=tipo_pag, cod_cliente=cod_cli_param)
                if cod_cli_param
                else ''
            )
            tipo_txt = 'Saída' if (id_.tipo_operacao or '') == '1' else 'Entrada'
            ws_nfe.append(
                [
                    nfe.id_nfe,
                    id_.numero or '',
                    id_.serie or '',
                    id_.chave_acesso or '',
                    id_.emissao.isoformat() if id_.emissao else '',
                    tipo_txt,
                    nfe.status or '',
                    nfe.empresa.cod_empresa if nfe.empresa else '',
                    nfe.filial.cod_filial if nfe.filial else '',
                    (nfe.filial.nome or '') if nfe.filial else '',
                    id_.natureza_operacao or '',
                    cond_sap,
                    'Sim' if nfe.tem_sap else 'Não',
                    nfe.sap_nome_tabela or '',
                ]
            )
            n_nfe += 1
        sheets_widths.append((ws_nfe, [9, 10, 6, 48, 20, 14, 14, 12, 10, 24, 36, 22, 12, 14]))

    if incluir_cte:
        ws_cte = wb.create_sheet('CTe')
        h_cte = [
            'id_cte',
            'Número',
            'Série',
            'Chave',
            'Emissão',
            'Empresa',
            'Filial',
            'Nome filial',
            'Chave no SAP',
            'Tabela SAP',
        ]
        ws_cte.append(h_cte)
        _style_header(ws_cte[1])
        n_cte = 0
        for cte in queryset_relatorio_cte(request, params).iterator(chunk_size=500):
            if n_cte >= MAX_ROWS:
                break
            id_ = cte.identificacao
            ws_cte.append(
                [
                    cte.id_cte,
                    id_.numero or '',
                    id_.serie or '',
                    id_.chave_acesso or '',
                    id_.emissao.isoformat() if id_.emissao else '',
                    cte.empresa.cod_empresa if cte.empresa else '',
                    cte.filial.cod_filial if cte.filial else '',
                    (cte.filial.nome or '') if cte.filial else '',
                    'Sim' if cte.tem_sap else 'Não',
                    cte.sap_nome_tabela or '',
                ]
            )
            n_cte += 1
        sheets_widths.append((ws_cte, [9, 10, 6, 48, 20, 12, 10, 24, 12, 14]))

    if incluir_nfse:
        ws_nfse = wb.create_sheet('NFS-e')
        h_nfse = [
            'id_nfse',
            'Número',
            'Chave',
            'Emissão',
            'Empresa',
            'Filial',
            'Nome filial',
            'Chave no SAP',
            'Tabela SAP',
        ]
        ws_nfse.append(h_nfse)
        _style_header(ws_nfse[1])
        n_nfse = 0
        for nfse in queryset_relatorio_nfse(request, params).iterator(chunk_size=500):
            if n_nfse >= MAX_ROWS:
                break
            id_ = nfse.identificacao
            ws_nfse.append(
                [
                    nfse.id_nfse,
                    id_.numero or '',
                    id_.chave or '',
                    id_.emissao.isoformat() if id_.emissao else '',
                    nfse.empresa.cod_empresa if nfse.empresa else '',
                    nfse.filial.cod_filial if nfse.filial else '',
                    (nfse.filial.nome or '') if nfse.filial else '',
                    'Sim' if nfse.tem_sap else 'Não',
                    nfse.sap_nome_tabela or '',
                ]
            )
            n_nfse += 1
        sheets_widths.append((ws_nfse, [9, 12, 48, 20, 12, 10, 24, 12, 14]))

    if incluir_sped:
        ws_sped = wb.create_sheet('SPED')
        h_sped = ['id_arquivo', 'Tipo', 'Tipo (texto)', 'Competência', 'Arquivo', 'Data carga', 'Empresa']
        ws_sped.append(h_sped)
        _style_header(ws_sped[1])
        n_sped = 0
        for it in list_relatorio_sped_items(request, params):
            if n_sped >= MAX_ROWS:
                break
            ws_sped.append(
                [
                    it.get('id_arquivo'),
                    it.get('tipo') or '',
                    it.get('tipo_display') or '',
                    it.get('competencia') or '',
                    it.get('nome_arquivo') or '',
                    it.get('data_carga') or '',
                    it.get('empresa') or '',
                ]
            )
            n_sped += 1
        sheets_widths.append((ws_sped, [12, 6, 18, 14, 48, 20, 12]))

    for ws, widths in sheets_widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = timezone.now().strftime('%Y%m%d_%H%M')
    if excel_tipo == 'todos':
        fname = f'relatorio_fiscal_{stamp}.xlsx'
    elif excel_tipo == 'importacao':
        fname = f'relatorio_fiscal_importacao_{stamp}.xlsx'
    else:
        fname = f'relatorio_fiscal_{excel_tipo}_{stamp}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _serialize_model(inst, exclude=None):
    """Converte um model instance em dict para JSON (datas e decimals já serializáveis)."""
    if inst is None:
        return None
    exclude = set(exclude or [])
    from django.forms.models import model_to_dict
    d = model_to_dict(inst, exclude=exclude)
    for k, v in list(d.items()):
        if hasattr(v, 'isoformat'):
            d[k] = v.isoformat() if v else None
        elif hasattr(v, '__float__') and not isinstance(v, (bool, int)):
            try:
                d[k] = float(v)
            except (TypeError, ValueError):
                pass
    return d


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_nfe_detalhe(request, id_nfe):
    """Detalhe completo da NFe para modal: cabeçalho, itens, total, cobrança/parcelas, pagamento, transporte, info adicionais."""
    empresas = relatorio_empresas_queryset(request)
    cod_empresas = list(empresas.values_list('cod_empresa', flat=True))
    cod_cliente = request.session.get('cod_cliente', None)
    qs_acesso = NFe.objects.filter(
        Q(empresa__cod_empresa__in=cod_empresas) |
        Q(empresa__isnull=True, gdfcliente__cod_cliente=cod_cliente)
    )
    nfe = get_object_or_404(
        qs_acesso.select_related(
            'identificacao', 'emitente', 'destinatario', 'empresa',
            'emitente__endereco', 'destinatario__endereco',
        ).prefetch_related(
            'identificacao__produtos__icms',
            'identificacao__produtos__pis',
            'identificacao__produtos__cofins',
            'identificacao__produtos__ipi',
            'identificacao__totalizacao',
            'identificacao__cobranca__parcelas',
            'identificacao__pagamento',
            'identificacao__transporte',
            'identificacao__informacoes_adicionais',
        ),
        id_nfe=id_nfe,
    )
    ide = nfe.identificacao
    cabecalho = {
        'identificacao': _serialize_model(ide),
        'emitente': _serialize_model(nfe.emitente),
        'destinatario': _serialize_model(nfe.destinatario),
        'nfe': _serialize_model(nfe, exclude=['identificacao', 'emitente', 'destinatario', 'empresa']),
        'empresa': nfe.empresa.cod_empresa if nfe.empresa else None,
    }
    if nfe.emitente and nfe.emitente.endereco:
        cabecalho['emitente_endereco'] = _serialize_model(nfe.emitente.endereco)
    if nfe.destinatario and nfe.destinatario.endereco:
        cabecalho['destinatario_endereco'] = _serialize_model(nfe.destinatario.endereco)
    itens = []
    for p in ide.produtos.all().order_by('numero_item'):
        item = _serialize_model(p, exclude=['nfe_serie'])
        item['icms'] = _serialize_model(p.icms, exclude=['produto']) if getattr(p, 'icms', None) else None
        item['pis'] = _serialize_model(p.pis, exclude=['produto']) if getattr(p, 'pis', None) else None
        item['cofins'] = _serialize_model(p.cofins, exclude=['produto']) if getattr(p, 'cofins', None) else None
        item['ipi'] = _serialize_model(p.ipi, exclude=['produto']) if getattr(p, 'ipi', None) else None
        itens.append(item)
    totalizacao = _serialize_model(ide.totalizacao) if hasattr(ide, 'totalizacao') and ide.totalizacao else None
    cobranca = None
    parcelas = []
    try:
        if getattr(ide, 'cobranca', None):
            cobranca = _serialize_model(ide.cobranca, exclude=['nfe_identificacao'])
            parcelas = [_serialize_model(parc, exclude=['nfe_cobranca']) for parc in ide.cobranca.parcelas.all().order_by('numero_parcela')]
    except Exception:
        cobranca = None
        parcelas = []
    pagamento = None
    try:
        if getattr(ide, 'pagamento', None):
            pagamento = _serialize_model(ide.pagamento, exclude=['nfe_identificacao'])
            # Descrição do tipo de pagamento: json/Tipo_pagamento.json, senão display do model
            pagamento['tipo_pagamento'] = (
                descricao_tipo_pagamento(ide.pagamento.meio_pagamento)
                or ide.pagamento.get_meio_pagamento_display()
            )
            if ide.pagamento.cartao_bandeira:
                pagamento['bandeira_cartao'] = ide.pagamento.get_cartao_bandeira_display()
            if ide.pagamento.pix_tipo_chave:
                pagamento['pix_tipo_chave_desc'] = ide.pagamento.get_pix_tipo_chave_display()
    except Exception:
        pagamento = None
    transporte = _serialize_model(ide.transporte) if hasattr(ide, 'transporte') and ide.transporte else None
    info_adic = _serialize_model(ide.informacoes_adicionais) if hasattr(ide, 'informacoes_adicionais') and ide.informacoes_adicionais else None
    return JsonResponse({
        'sucesso': True,
        'cabecalho': cabecalho,
        'itens': itens,
        'totalizacao': totalizacao,
        'cobranca': cobranca,
        'parcelas': parcelas,
        'pagamento': pagamento,
        'transporte': transporte,
        'informacoes_adicionais': info_adic,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_cte_detalhe(request, id_cte):
    """Detalhe completo do CTe para modal: cabeçalho, valor, transporte, carga, serviço, veículo, motorista, percurso, fiscal."""
    empresas = relatorio_empresas_queryset(request)
    cod_empresas = list(empresas.values_list('cod_empresa', flat=True))
    cod_cliente = request.session.get('cod_cliente', None)
    qs_acesso = CTe.objects.filter(
        Q(empresa__cod_empresa__in=cod_empresas) |
        Q(empresa__isnull=True, gdfcliente__cod_cliente=cod_cliente)
    )
    cte = get_object_or_404(
        qs_acesso.select_related(
            'identificacao', 'emitente', 'destinatario', 'empresa',
            'emitente__endereco', 'destinatario__endereco',
        ),
        id_cte=id_cte,
    )
    ide = cte.identificacao
    cabecalho = {
        'identificacao': _serialize_model(ide),
        'emitente': _serialize_model(cte.emitente),
        'destinatario': _serialize_model(cte.destinatario),
        'cte': _serialize_model(cte, exclude=['identificacao', 'emitente', 'destinatario', 'empresa']),
        'empresa': cte.empresa.cod_empresa if cte.empresa else None,
    }
    if cte.emitente and cte.emitente.endereco:
        cabecalho['emitente_endereco'] = _serialize_model(cte.emitente.endereco)
    if cte.destinatario and cte.destinatario.endereco:
        cabecalho['destinatario_endereco'] = _serialize_model(cte.destinatario.endereco)
    valor = _serialize_model(ide.valor) if hasattr(ide, 'valor') and ide.valor else None
    transporte = _serialize_model(ide.transporte) if hasattr(ide, 'transporte') and ide.transporte else None
    carga = _serialize_model(ide.carga) if hasattr(ide, 'carga') and ide.carga else None
    servico = _serialize_model(ide.servico) if hasattr(ide, 'servico') and ide.servico else None
    veiculo = _serialize_model(ide.veiculo) if hasattr(ide, 'veiculo') and ide.veiculo else None
    motorista = _serialize_model(ide.motorista) if hasattr(ide, 'motorista') and ide.motorista else None
    percurso = _serialize_model(ide.percurso) if hasattr(ide, 'percurso') and ide.percurso else None
    fiscal = _serialize_model(ide.fiscal) if hasattr(ide, 'fiscal') and ide.fiscal else None
    return JsonResponse({
        'sucesso': True,
        'cabecalho': cabecalho,
        'valor': valor,
        'transporte': transporte,
        'carga': carga,
        'servico': servico,
        'veiculo': veiculo,
        'motorista': motorista,
        'percurso': percurso,
        'fiscal': fiscal,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaXml', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_nfse_detalhe(request, id_nfse):
    """Detalhe completo da NFSe para modal: cabeçalho, prestador, tomador, serviços, RPS, retenção, pagamento."""
    empresas = relatorio_empresas_queryset(request)
    cod_empresas = list(empresas.values_list('cod_empresa', flat=True))
    cod_cliente = request.session.get('cod_cliente', None)
    qs_acesso = NFSe.objects.filter(
        Q(empresa__cod_empresa__in=cod_empresas) |
        Q(empresa__isnull=True, gdfcliente__cod_cliente=cod_cliente)
    )
    nfse = get_object_or_404(
        qs_acesso.select_related(
            'identificacao', 'prestador', 'tomador', 'empresa',
            'prestador__endereco', 'tomador__endereco',
        ).prefetch_related(
            'identificacao__servicos',
            'identificacao__rps_list',
        ),
        id_nfse=id_nfse,
    )
    ide = nfse.identificacao
    cabecalho = {
        'identificacao': _serialize_model(ide),
        'prestador': _serialize_model(nfse.prestador),
        'tomador': _serialize_model(nfse.tomador),
        'nfse': _serialize_model(nfse, exclude=['identificacao', 'prestador', 'tomador', 'empresa']),
        'empresa': nfse.empresa.cod_empresa if nfse.empresa else None,
    }
    if nfse.prestador and nfse.prestador.endereco:
        cabecalho['prestador_endereco'] = _serialize_model(nfse.prestador.endereco)
    if nfse.tomador and nfse.tomador.endereco:
        cabecalho['tomador_endereco'] = _serialize_model(nfse.tomador.endereco)
    servicos = [_serialize_model(s) for s in ide.servicos.all()]
    rps_list = [_serialize_model(r) for r in ide.rps_list.all()]
    retencao = _serialize_model(ide.retencao) if hasattr(ide, 'retencao') and ide.retencao else None
    pagamento = _serialize_model(ide.pagamento) if hasattr(ide, 'pagamento') and ide.pagamento else None
    return JsonResponse({
        'sucesso': True,
        'cabecalho': cabecalho,
        'servicos': servicos,
        'rps': rps_list,
        'retencao': retencao,
        'pagamento': pagamento,
    }, status=200)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Pro_CargaSped', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_relatorio_sped_detalhe(request, tipo, id_arquivo):
    """Detalhe do arquivo SPED: cabeçalho e registros. tipo: F=Fiscal, C=Contribuição."""
    from app.db_GDF.sped_fiscal.models import (
        SpedFiscalArquivo,
        SpedFiscalReg_0000, SpedFiscalReg_0001, SpedFiscalReg_0005, SpedFiscalReg_0150,
        SpedFiscalReg_0190, SpedFiscalReg_0200, SpedFiscalReg_C001, SpedFiscalReg_C100,
        SpedFiscalReg_C170, SpedFiscalReg_C190, SpedFiscalReg_D100, SpedFiscalRegistro,
    )
    from app.db_GDF.sped_contribuicao.models import (
        SpedContribuicaoArquivo,
        SpedContribuicaoReg_0000, SpedContribuicaoReg_0001, SpedContribuicaoReg_0005,
        SpedContribuicaoReg_0150, SpedContribuicaoReg_0190, SpedContribuicaoReg_0200,
        SpedContribuicaoReg_C001, SpedContribuicaoReg_C100, SpedContribuicaoReg_C170,
        SpedContribuicaoReg_C190, SpedContribuicaoReg_D100, SpedContribuicaoRegistro,
    )
    from decimal import Decimal

    if tipo not in ('F', 'C'):
        return JsonResponse({'sucesso': False, 'mensagem': 'Tipo SPED inválido'}, status=400)
    Arquivo = SpedFiscalArquivo if tipo == 'F' else SpedContribuicaoArquivo

    empresas = relatorio_empresas_queryset(request)
    cod_empresas = list(empresas.values_list('cod_empresa', flat=True))
    cod_cliente = request.session.get('cod_cliente', None)
    q_detalhe = Q(empresa__cod_empresa__in=cod_empresas)
    if cod_cliente:
        q_detalhe |= Q(empresa__isnull=True, gdfcliente__cod_cliente=cod_cliente)
    qs_arquivo = Arquivo.objects.filter(q_detalhe).prefetch_related(
        'reg_0000', 'reg_0001', 'reg_0005', 'reg_0150', 'reg_0190', 'reg_0200',
        'reg_c001', 'reg_c100', 'reg_c170', 'reg_c190', 'reg_d100', 'registros',
    ).select_related('empresa')
    arq = get_object_or_404(qs_arquivo, id_arquivo=id_arquivo)
    cabecalho = _serialize_model(arq)
    if cabecalho:
        cabecalho['empresa'] = arq.empresa.cod_empresa if arq.empresa else None
        cabecalho['tipo'] = tipo
        cabecalho['tipo_display'] = 'Fiscal' if tipo == 'F' else 'Contribuição'

    def _serialize_dec(v):
        return float(v) if v is not None and isinstance(v, Decimal) else v

    reg_0000 = [{'linha': r.linha, 'cod_ver': r.cod_ver, 'dt_ini': str(r.dt_ini) if r.dt_ini else None, 'dt_fin': str(r.dt_fin) if r.dt_fin else None, 'nome': r.nome, 'cnpj': r.cnpj} for r in arq.reg_0000.all()[:50]]
    reg_0001 = [{'linha': r.linha, 'ind_mov': r.ind_mov} for r in arq.reg_0001.all()[:50]]
    reg_0005 = [{'linha': r.linha, 'fantasia': r.fantasia, 'end': r.end, 'bairro': r.bairro, 'email': r.email} for r in arq.reg_0005.all()[:20]]
    reg_0150 = [{'linha': r.linha, 'cod_part': r.cod_part, 'nome': r.nome, 'cnpj': r.cnpj, 'end': r.end} for r in arq.reg_0150.all()[:100]]
    reg_0190 = [{'linha': r.linha, 'unid': r.unid, 'descr': r.descr} for r in arq.reg_0190.all()[:50]]
    reg_0200 = [{'linha': r.linha, 'cod_item': r.cod_item, 'descr_item': r.descr_item, 'unid_inv': r.unid_inv, 'cod_ncm': r.cod_ncm} for r in arq.reg_0200.all()[:200]]
    reg_c001 = [{'linha': r.linha, 'ind_mov': r.ind_mov} for r in arq.reg_c001.all()[:20]]
    # C100: documento fiscal com impostos (ICMS, PIS, COFINS, IPI)
    reg_c100 = [{
        'linha': r.linha, 'chv_nfe': r.chv_nfe, 'dt_doc': str(r.dt_doc) if r.dt_doc else None,
        'vl_doc': _serialize_dec(r.vl_doc), 'num_doc': r.num_doc, 'ser': r.ser,
        'ind_oper': r.ind_oper, 'ind_emit': r.ind_emit,
        'vl_bc_icms': _serialize_dec(r.vl_bc_icms), 'vl_icms': _serialize_dec(r.vl_icms),
        'vl_bc_icms_st': _serialize_dec(r.vl_bc_icms_st), 'vl_icms_st': _serialize_dec(r.vl_icms_st),
        'vl_ipi': _serialize_dec(r.vl_ipi), 'vl_pis': _serialize_dec(r.vl_pis), 'vl_cofins': _serialize_dec(r.vl_cofins),
    } for r in arq.reg_c100.all()[:200]]
    # C170: itens com impostos por produto (ICMS, PIS, COFINS)
    reg_c170 = [{
        'linha': r.linha, 'num_item': r.num_item, 'cod_item': r.cod_item, 'descr_compl': r.descr_compl,
        'qtd': _serialize_dec(r.qtd), 'unid': r.unid, 'vl_item': _serialize_dec(r.vl_item), 'vl_desc': _serialize_dec(r.vl_desc),
        'cst_icms': r.cst_icms, 'cfop': r.cfop,
        'vl_bc_icms': _serialize_dec(r.vl_bc_icms), 'aliq_icms': _serialize_dec(r.aliq_icms), 'vl_icms': _serialize_dec(r.vl_icms),
        'vl_bc_icms_st': _serialize_dec(r.vl_bc_icms_st), 'aliq_st': _serialize_dec(r.aliq_st), 'vl_icms_st': _serialize_dec(r.vl_icms_st),
        'cst_pis': r.cst_pis, 'vl_bc_pis': _serialize_dec(r.vl_bc_pis), 'aliq_pis': _serialize_dec(r.aliq_pis), 'vl_pis': _serialize_dec(r.vl_pis),
        'cst_cofins': r.cst_cofins, 'vl_bc_cofins': _serialize_dec(r.vl_bc_cofins), 'aliq_cofins': _serialize_dec(r.aliq_cofins), 'vl_cofins': _serialize_dec(r.vl_cofins),
    } for r in arq.reg_c170.all()[:500]]
    # C190: Fiscal=analítico ICMS (CST/CFOP); Contribuição=consolidação por item (COD_ITEM, PIS/COFINS)
    reg_c190 = [{
        'linha': r.linha, 'cod_item': getattr(r, 'cod_item', None), 'cst_icms': r.cst_icms, 'cfop': r.cfop,
        'cst_pis': getattr(r, 'cst_pis', None), 'vl_bc_pis': _serialize_dec(getattr(r, 'vl_bc_pis', None)),
        'vl_pis': _serialize_dec(getattr(r, 'vl_pis', None)), 'cst_cofins': getattr(r, 'cst_cofins', None),
        'vl_bc_cofins': _serialize_dec(getattr(r, 'vl_bc_cofins', None)), 'vl_cofins': _serialize_dec(getattr(r, 'vl_cofins', None)),
        'vl_opr': _serialize_dec(r.vl_opr), 'vl_bc_icms': _serialize_dec(r.vl_bc_icms), 'aliq_icms': _serialize_dec(r.aliq_icms),
        'vl_icms': _serialize_dec(r.vl_icms), 'vl_bc_icms_st': _serialize_dec(r.vl_bc_icms_st), 'vl_icms_st': _serialize_dec(r.vl_icms_st),
        'vl_red_bc': _serialize_dec(r.vl_red_bc), 'vl_ipi': _serialize_dec(r.vl_ipi),
    } for r in arq.reg_c190.all()[:300]]
    reg_d100 = [{'linha': r.linha, 'chv_cte': r.chv_cte, 'dt_doc': str(r.dt_doc) if r.dt_doc else None, 'vl_doc': _serialize_dec(r.vl_doc)} for r in arq.reg_d100.all()[:200]]
    registros = [{'registro': r.registro, 'linha': r.linha, 'campos': r.campos, 'conteudo': (r.conteudo or '')[:500]} for r in arq.registros.all()[:300]]
    registros_fiscal = []
    registros_contribuicao = []

    return JsonResponse({
        'sucesso': True,
        'cabecalho': cabecalho,
        'reg_0000': reg_0000,
        'reg_0001': reg_0001,
        'reg_0005': reg_0005,
        'reg_0150': reg_0150,
        'reg_0190': reg_0190,
        'reg_0200': reg_0200,
        'reg_c001': reg_c001,
        'reg_c100': reg_c100,
        'reg_c170': reg_c170,
        'reg_c190': reg_c190,
        'reg_d100': reg_d100,
        'registros': registros,
        'registros_fiscal': registros_fiscal,
        'registros_contribuicao': registros_contribuicao,
    }, status=200)


@login_required(login_url='Login')
@require_http_methods(["GET"])
def fn_view_Relatorio_Fiscal(request):
    """
    Rota legada: o relatório fiscal integrou-se à Carga de XML (NFe, CT-e, NFS) e Carga SPED (SPED).
    Redireciona para a Carga de XML.
    """
    return redirect("Pro_CargaXml")


# -------------------------------------------------------------------------
# Ferramentas – subsolução Reproc_Painel (painel de reprocessamento, confronto SPED x NFe)
# -------------------------------------------------------------------------
@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel')
def fn_view_Reprocessamento(request):
    """Legado: redireciona para o Painel."""
    return redirect('Reproc_Painel')


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel')
def fn_view_Reprocessamento_Painel(request):
    """Painel de Reprocessamento: confronto SPED x NFe por empresa, divergências e reprocessamento controlado."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        context = {'cod_cliente': None, 'empresas': [], 'tipo_pagamento_desc': TIPO_PAGAMENTO_DESC}
        return render(request, 'Ferramentas/painel.html', context)
    empresas = list(
        Empresa.objects.filter(gdfcliente_id=cod_cliente).values('cod_empresa', 'razao', 'fantasia').order_by('razao')
    )
    context = {
        'cod_cliente': cod_cliente,
        'empresas': empresas,
        'tipo_pagamento_desc': TIPO_PAGAMENTO_DESC,
    }
    return render(request, 'Ferramentas/painel.html', context)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_lotes(request):
    """Lista lotes de reprocessamento do cliente (filtros: empresa, competência, status)."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    if not cod_empresas:
        return JsonResponse({'sucesso': True, 'lotes': [], 'total': 0})

    qs = ReprocessamentoLote.objects.filter(empresa_id__in=cod_empresas).select_related('empresa')
    cod_empresa = request.GET.get('empresa') or request.GET.get('grupo')
    if cod_empresa and cod_empresa in cod_empresas:
        qs = qs.filter(empresa_id=cod_empresa)
    competencia = request.GET.get('competencia')
    if competencia:
        competencia = competencia.strip()
        if len(competencia) == 7 and competencia[4] == '-':  # YYYY-MM
            from datetime import datetime
            try:
                dt_comp = datetime.strptime(competencia + '-01', '%Y-%m-%d').date()
                qs = qs.filter(competencia=dt_comp)
            except ValueError:
                pass
        else:
            try:
                from datetime import datetime
                dt_comp = datetime.strptime(competencia, '%Y-%m-%d').date()
                qs = qs.filter(competencia=dt_comp)
            except ValueError:
                pass
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)
    criado_de = request.GET.get('criado_de')  # YYYY-MM-DD
    criado_ate = request.GET.get('criado_ate')
    if criado_de:
        try:
            from datetime import datetime
            qs = qs.filter(data_criacao__date__gte=datetime.strptime(criado_de, '%Y-%m-%d').date())
        except ValueError:
            pass
    if criado_ate:
        try:
            from datetime import datetime
            qs = qs.filter(data_criacao__date__lte=datetime.strptime(criado_ate, '%Y-%m-%d').date())
        except ValueError:
            pass
    qs = qs.order_by('-data_criacao')[:500]
    lotes = [
        {
            'id_lote': l.id_lote,
            'cod_empresa': getattr(l.empresa, 'cod_empresa', l.empresa_id) if l.empresa_id else None,
            'empresa_razao': getattr(l.empresa, 'razao', None) if l.empresa_id else None,
            'empresa_fantasia': getattr(l.empresa, 'fantasia', None) if l.empresa_id else None,
            'competencia': str(l.competencia),
            'competencia_mes': l.competencia.strftime('%Y-%m') if l.competencia else None,
            'id_arquivo_sped': l.id_arquivo_sped,
            'total_nfe_esperado': l.total_nfe_esperado,
            'total_nfe_encontrado': l.total_nfe_encontrado,
            'total_divergencias': l.total_divergencias,
            'status': l.status,
            'mensagem_erro': l.mensagem_erro,
            'usuario_criacao': l.usuario_criacao,
            'data_inicio': l.data_inicio.isoformat() if l.data_inicio else None,
            'data_fim': l.data_fim.isoformat() if l.data_fim else None,
            'data_criacao': l.data_criacao.isoformat(),
        }
        for l in qs
    ]
    return JsonResponse({'sucesso': True, 'lotes': lotes, 'total': len(lotes)})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_divergencias(request, id_lote):
    """Lista divergências de um lote."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    lote = get_object_or_404(ReprocessamentoLote, id_lote=id_lote, empresa_id__in=cod_empresas)
    total_divs = Divergencia.objects.filter(lote=lote).count()
    divs = Divergencia.objects.filter(lote=lote).order_by('tipo', 'chave_nfe', '-data_criacao')[:5000]
    lista = [
        {
            'id_divergencia': d.id_divergencia,
            'cod_empresa': d.cod_empresa,
            'tipo': d.tipo,
            'status': d.status,
            'chave_nfe': d.chave_nfe,
            'numero_nfe': d.numero_nfe,
            'serie_nfe': d.serie_nfe,
            'descricao': d.descricao,
            'valor_esperado': str(d.valor_esperado) if d.valor_esperado is not None else None,
            'valor_encontrado': str(d.valor_encontrado) if d.valor_encontrado is not None else None,
            'registro_sped': d.registro_sped,
            'linha_sped': d.linha_sped,
            'id_nfe': d.id_nfe,
            'detalhe_json': d.detalhe_json,
            'data_reprocessamento': d.data_reprocessamento.isoformat() if d.data_reprocessamento else None,
            'usuario_reprocessamento': d.usuario_reprocessamento,
            'data_criacao': d.data_criacao.isoformat(),
            'data_atualizacao': d.data_atualizacao.isoformat() if d.data_atualizacao else None,
        }
        for d in divs
    ]
    return JsonResponse({
        'sucesso': True,
        'divergencias': lista,
        'lote_id': lote.id_lote,
        'total': total_divs,
    })


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_confronto(request):
    """Dispara confronto SPED x NFe para uma empresa e competência (mês)."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    if not cod_empresas:
        return JsonResponse({'sucesso': False, 'mensagem': 'Nenhuma empresa vinculada ao cliente.'}, status=400)

    data = json.loads(request.body) if request.body else {}
    competencia = data.get('competencia')
    cod_empresa = (data.get('cod_empresa') or data.get('empresa') or '').strip()
    if not cod_empresa:
        return JsonResponse({'sucesso': False, 'mensagem': 'Empresa obrigatória (cod_empresa).'}, status=400)
    if cod_empresa not in cod_empresas:
        return JsonResponse({'sucesso': False, 'mensagem': 'Empresa não permitida para este cliente.'}, status=403)
    if not competencia:
        return JsonResponse({'sucesso': False, 'mensagem': 'Competência obrigatória (mês: YYYY-MM).'}, status=400)

    from datetime import datetime
    competencia = competencia.strip()
    try:
        if len(competencia) == 7 and competencia[4] == '-':
            dt = datetime.strptime(competencia + '-01', '%Y-%m-%d').date()
        else:
            dt = datetime.strptime(competencia, '%Y-%m-%d').date()
            dt = dt.replace(day=1)
    except ValueError:
        return JsonResponse({'sucesso': False, 'mensagem': 'Competência inválida. Use YYYY-MM (mês).'}, status=400)

    usuario = getattr(request.user, 'username', '') or str(request.user)
    lote = ReprocessamentoLote.objects.create(
        empresa_id=cod_empresa,
        competencia=dt,
        status='PENDENTE',
        usuario_criacao=usuario,
    )
    job = ReprocessamentoJob.objects.create(
        tipo='CONFRONTO',
        status='AGUARDANDO',
        id_lote=lote.id_lote,
        usuario=usuario,
    )
    lote.status = 'EM_CONFRONTO'
    lote.data_inicio = timezone.now()
    lote.save(update_fields=['status', 'data_inicio'])
    job.status = 'EM_EXECUCAO'
    job.data_inicio = timezone.now()
    job.save(update_fields=['status', 'data_inicio'])
    try:
        from app.classes.Reprocessamento import confrontar_sped_nfe
        confrontar_sped_nfe(lote.id_lote, cod_empresa, dt)
    except Exception:
        lote.total_nfe_esperado = 0
        lote.total_nfe_encontrado = 0
        lote.total_divergencias = 0
        lote.status = 'CONCLUIDO'
        lote.data_fim = timezone.now()
        lote.save(update_fields=['total_nfe_esperado', 'total_nfe_encontrado', 'total_divergencias', 'status', 'data_fim'])
        job.status = 'CONCLUIDO'
        job.data_fim = timezone.now()
        job.total_processados = 0
        job.save(update_fields=['status', 'data_fim', 'total_processados'])

    return JsonResponse({
        'sucesso': True,
        'id_lote': lote.id_lote,
        'ids_lotes': [lote.id_lote],
        'total_lotes': 1,
        'mensagem': 'Confronto iniciado. Consulte o painel para acompanhar.',
    })


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_divergencia_detalhe(request, id_divergencia):
    """Retorna detalhe completo da divergência: resumo, cabeçalho NFe/SPED, itens, impostos e confrontos realizados."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    div = get_object_or_404(Divergencia, id_divergencia=id_divergencia)
    if (div.lote.empresa_id or '') not in cod_empresas:
        return JsonResponse({'sucesso': False, 'mensagem': 'Divergência não pertence ao cliente'}, status=403)

    from decimal import Decimal

    payload = {
        'divergencia': {
            'id_divergencia': div.id_divergencia,
            'cod_empresa': div.cod_empresa,
            'tipo': div.tipo,
            'status': div.status,
            'chave_nfe': div.chave_nfe,
            'numero_nfe': div.numero_nfe,
            'serie_nfe': div.serie_nfe,
            'descricao': div.descricao,
            'valor_esperado': str(div.valor_esperado) if div.valor_esperado is not None else None,
            'valor_encontrado': str(div.valor_encontrado) if div.valor_encontrado is not None else None,
            'registro_sped': div.registro_sped,
            'linha_sped': div.linha_sped,
            'id_nfe': div.id_nfe,
            'detalhe_json': div.detalhe_json,
            'data_criacao': div.data_criacao.isoformat() if div.data_criacao else None,
            'data_reprocessamento': div.data_reprocessamento.isoformat() if div.data_reprocessamento else None,
            'usuario_reprocessamento': div.usuario_reprocessamento,
        },
        'nfe': None,
        'sped': None,
        'confrontos': [],
    }

    # NFe: cabeçalho, itens, impostos (quando id_nfe existe)
    if div.id_nfe:
        try:
            nfe = NFe.objects.filter(id_nfe=div.id_nfe).select_related(
                'identificacao', 'emitente', 'destinatario'
            ).prefetch_related(
                'identificacao__produtos',
                'identificacao__totalizacao',
            ).first()
            if nfe:
                ide = nfe.identificacao
                tot = getattr(ide, 'totalizacao', None)
                produtos = list(ide.produtos.order_by('numero_item').values(
                    'id_produto', 'numero_item', 'descricao', 'cfop', 'ncm', 'quantidade',
                    'valor_unitario', 'valor_total', 'unidade'
                ))
                for p in produtos:
                    p['quantidade'] = str(p['quantidade']) if p.get('quantidade') is not None else None
                    p['valor_unitario'] = str(p['valor_unitario']) if p.get('valor_unitario') is not None else None
                    p['valor_total'] = str(p['valor_total']) if p.get('valor_total') is not None else None
                    try:
                        prod_obj = ide.produtos.get(id_produto=p['id_produto'])
                        try:
                            icms = prod_obj.icms
                            p['icms'] = {'cst': icms.cst, 'valor': str(icms.valor_icms or 0)}
                        except Exception:
                            p['icms'] = None
                        try:
                            pis = prod_obj.pis
                            p['pis'] = {'cst': pis.cst, 'valor': str(pis.valor_pis or 0)}
                        except Exception:
                            p['pis'] = None
                        try:
                            cofins = prod_obj.cofins
                            p['cofins'] = {'cst': cofins.cst, 'valor': str(cofins.valor_cofins or 0)}
                        except Exception:
                            p['cofins'] = None
                    except Exception:
                        pass

                payload['nfe'] = {
                    'cabeçalho': {
                        'numero': ide.numero,
                        'serie': ide.serie,
                        'chave_acesso': ide.chave_acesso,
                        'emissao': ide.emissao.isoformat() if ide.emissao else None,
                        'natureza_operacao': ide.natureza_operacao,
                        'modelo': ide.modelo,
                        'emitente': nfe.emitente.razao_social if nfe.emitente else None,
                        'destinatario': nfe.destinatario.razao_social if nfe.destinatario else None,
                    },
                    'totalizacao': {
                        'valor_subtotal_produtos': str(tot.valor_subtotal_produtos) if tot else None,
                        'valor_frete': str(tot.valor_frete) if tot and tot.valor_frete else None,
                        'valor_desconto': str(tot.valor_desconto) if tot and tot.valor_desconto else None,
                        'valor_base_icms': str(tot.valor_base_icms) if tot and tot.valor_base_icms else None,
                        'valor_icms': str(tot.valor_icms) if tot and tot.valor_icms else None,
                        'valor_icms_st': str(tot.valor_icms_st) if tot and tot.valor_icms_st else None,
                        'valor_pis': str(tot.valor_pis) if tot and tot.valor_pis else None,
                        'valor_cofins': str(tot.valor_cofins) if tot and tot.valor_cofins else None,
                        'valor_total_nfe': str(tot.valor_total_nfe) if tot else None,
                    } if tot else None,
                    'itens': produtos,
                }
        except Exception:
            payload['nfe'] = None

    # SPED: C100 e C170 (busca em sped_fiscal e sped_contribuicao)
    if div.chave_nfe:
        try:
            from app.db_GDF.sped_contribuicao.models import SpedContribuicaoReg_C100, SpedContribuicaoReg_C170
            c100 = SpedFiscalReg_C100.objects.filter(chv_nfe=div.chave_nfe).select_related('arquivo').first()
            if not c100:
                c100 = SpedContribuicaoReg_C100.objects.filter(chv_nfe=div.chave_nfe).select_related('arquivo').first()
            if c100:
                Reg_C170 = SpedFiscalReg_C170 if isinstance(c100, SpedFiscalReg_C100) else SpedContribuicaoReg_C170
                c170_list = list(Reg_C170.objects.filter(c100=c100).order_by('linha', 'num_item').values(
                    'num_item', 'cod_item', 'descr_compl', 'cfop', 'qtd', 'unid',
                    'vl_item', 'vl_desc', 'cst_icms', 'vl_bc_icms', 'aliq_icms', 'vl_icms',
                    'cst_pis', 'vl_bc_pis', 'aliq_pis', 'vl_pis',
                    'cst_cofins', 'vl_bc_cofins', 'aliq_cofins', 'vl_cofins',
                ))
                for c in c170_list:
                    for k, v in list(c.items()):
                        if isinstance(v, Decimal):
                            c[k] = str(v)

                payload['sped'] = {
                    'cabeçalho': {
                        'chv_nfe': c100.chv_nfe,
                        'num_doc': c100.num_doc,
                        'ser': c100.ser,
                        'dt_doc': str(c100.dt_doc) if c100.dt_doc else None,
                        'vl_doc': str(c100.vl_doc) if c100.vl_doc else None,
                        'vl_bc_icms': str(c100.vl_bc_icms) if c100.vl_bc_icms else None,
                        'vl_icms': str(c100.vl_icms) if c100.vl_icms else None,
                        'vl_icms_st': str(c100.vl_icms_st) if c100.vl_icms_st else None,
                        'vl_pis': str(c100.vl_pis) if c100.vl_pis else None,
                        'vl_cofins': str(c100.vl_cofins) if c100.vl_cofins else None,
                    },
                    'itens': c170_list,
                }
        except Exception:
            payload['sped'] = None

    # Confrontos realizados e status
    nfe_h = payload.get('nfe', {}).get('cabeçalho') if payload.get('nfe') else None
    sped_h = payload.get('sped', {}).get('cabeçalho') if payload.get('sped') else None
    nfe_tot = payload.get('nfe', {}).get('totalizacao') if payload.get('nfe') else None

    payload['confrontos'] = [
        {
            'tipo': 'Estrutural (documento)',
            'descricao': 'Documento presente no SPED e na NF-e',
            'status': 'DIVERGÊNCIA' if div.tipo in ('NFE_AUSENTE_SPED', 'SPED_AUSENTE_NFE') else 'OK',
            'detalhe': div.get_tipo_display() if hasattr(div, 'get_tipo_display') else div.tipo,
        },
        {
            'tipo': 'Valor do documento',
            'descricao': 'Valor total do documento',
            'status': 'OK' if (div.valor_esperado is None and div.valor_encontrado is None) or (
                div.valor_esperado == div.valor_encontrado
            ) else ('DIVERGÊNCIA' if div.valor_esperado or div.valor_encontrado else 'N/A'),
            'detalhe': f'SPED: {div.valor_esperado} | NFe: {div.valor_encontrado}' if div.valor_esperado or div.valor_encontrado else None,
        },
        {
            'tipo': 'Cabeçalho (data emissão)',
            'descricao': 'Data de emissão do documento',
            'status': 'OK' if (nfe_h and sped_h and nfe_h.get('emissao') and sped_h.get('dt_doc') and
                str(nfe_h.get('emissao', '')[:10]) == str(sped_h.get('dt_doc', ''))) else ('DIVERGÊNCIA' if (nfe_h and sped_h) else 'N/A'),
            'detalhe': f'NFe: {nfe_h.get("emissao", "")[:10] if nfe_h else "-"} | SPED: {sped_h.get("dt_doc", "-") if sped_h else "-"}' if (nfe_h or sped_h) else None,
        },
        {
            'tipo': 'Itens',
            'descricao': 'Quantidade e valores dos itens',
            'status': 'N/A' if not (payload.get('nfe') and payload.get('sped')) else 'OK',
            'detalhe': f'NFe: {len(payload["nfe"].get("itens", []))} itens | SPED: {len(payload["sped"].get("itens", []))} itens' if (payload.get('nfe') and payload.get('sped')) else None,
        },
        {
            'tipo': 'Impostos (ICMS, PIS, COFINS)',
            'descricao': 'Valores de impostos do documento',
            'status': 'OK' if (nfe_tot and sped_h and
                (not div.valor_esperado or not div.valor_encontrado or div.valor_esperado == div.valor_encontrado)) else ('DIVERGÊNCIA' if (nfe_tot and sped_h) else 'N/A'),
            'detalhe': f'NFe ICMS: {nfe_tot.get("valor_icms", "-") if nfe_tot else "-"} | SPED ICMS: {sped_h.get("vl_icms", "-") if sped_h else "-"}' if (nfe_tot or sped_h) else None,
        },
    ]

    return JsonResponse({'sucesso': True, 'detalhe': payload})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_reprocessar_divergencia(request, id_divergencia):
    """Marca divergência como resolvida após reprocessamento."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    div = get_object_or_404(Divergencia, id_divergencia=id_divergencia)
    if (div.lote.empresa_id or '') not in cod_empresas:
        return JsonResponse({'sucesso': False, 'mensagem': 'Divergência não pertence ao cliente'}, status=403)
    usuario = getattr(request.user, 'username', '') or str(request.user)
    div.status = 'RESOLVIDA'
    div.data_reprocessamento = timezone.now()
    div.usuario_reprocessamento = usuario
    div.save(update_fields=['status', 'data_reprocessamento', 'usuario_reprocessamento'])
    return JsonResponse({'sucesso': True, 'mensagem': 'Divergência marcada como resolvida.'})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_condicoes_gerar(request, id_lote):
    """Gera/atualiza registros de condição de pagamento para o lote (chaves 44 + condição NFe)."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    lote = get_object_or_404(ReprocessamentoLote, id_lote=id_lote, empresa_id__in=cod_empresas)
    try:
        from app.classes.Reprocessamento import gerar_condicoes_pagamento_lote
        criados, atualizados = gerar_condicoes_pagamento_lote(lote.id_lote)
        return JsonResponse({
            'sucesso': True,
            'criados': criados,
            'atualizados': atualizados,
            'mensagem': f'Gerado: {criados} novos, {atualizados} atualizados.',
        })
    except Exception as e:
        return JsonResponse({'sucesso': False, 'mensagem': str(e)[:500]}, status=500)


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_condicoes_listar(request, id_lote):
    """Lista condições de pagamento do lote (chave, condição NFe, SAP, retorno SAP, status)."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    lote = get_object_or_404(ReprocessamentoLote, id_lote=id_lote, empresa_id__in=cod_empresas)
    qs = CondicaoPagamentoLote.objects.filter(lote=lote).order_by('chave_nfe')
    lista = [
        {
            'id_reg': c.id_reg,
            'cod_empresa': c.cod_empresa,
            'chave_nfe': c.chave_nfe,
            'numero_nfe': c.numero_nfe,
            'serie_nfe': c.serie_nfe,
            'condicao_pagamento_nfe': c.condicao_pagamento_nfe,
            'condicao_pagamento_sap': c.condicao_pagamento_sap,
            'tipo_pagamento': c.tipo_pagamento or '',
            'status': c.status,
            'data_criacao': c.data_criacao.isoformat() if c.data_criacao else None,
            'data_atualizacao': c.data_atualizacao.isoformat() if c.data_atualizacao else None,
        }
        for c in qs
    ]
    return JsonResponse({'sucesso': True, 'condicoes': lista, 'total': len(lista)})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_condicoes_atualizar_retorno(request, id_lote):
    """
    Atualiza condicao_pagamento_sap e status (P/E/S/U/I) por chave.
    Body: { "itens": [ { "chave_nfe": "44...", "condicao_sap_retorno": "Z001" }, ... ] }
    """
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    lote = get_object_or_404(ReprocessamentoLote, id_lote=id_lote, empresa_id__in=cod_empresas)
    data = json.loads(request.body) if request.body else {}
    itens = data.get('itens') or data.get('retornos') or []
    if not itens:
        return JsonResponse({'sucesso': False, 'mensagem': 'Envie "itens" com chave_nfe e condicao_sap_retorno.'}, status=400)
    atualizados = 0
    chaves_lote = set(
        CondicaoPagamentoLote.objects.filter(lote=lote).values_list('chave_nfe', flat=True)
    )
    for item in itens:
        chave = item.get('chave_nfe') or item.get('chave')
        retorno = item.get('condicao_sap_retorno') or item.get('condicao_sap')
        if not chave or chave not in chaves_lote:
            continue
        n = CondicaoPagamentoLote.objects.filter(lote=lote, chave_nfe=chave).update(
            condicao_pagamento_sap=retorno or '',
            status=item.get('status', 'S'),
        )
        atualizados += n
    return JsonResponse({'sucesso': True, 'atualizados': atualizados, 'mensagem': f'{atualizados} registro(s) atualizado(s).'})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_condicoes_enviar_sap(request, id_lote):
    """
    Chama RFC e envia as condições de pagamento do lote ao SAP.
    Atualiza a tabela com a condição retornada pelo SAP para cada chave e status (U/I/S).
    """
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    cod_empresas = reprocessamento_empresas_cliente(cod_cliente)
    qs_lote = ReprocessamentoLote.objects.filter(empresa_id__in=cod_empresas).select_related('empresa')
    lote = get_object_or_404(qs_lote, id_lote=id_lote)
    condicoes = list(
        CondicaoPagamentoLote.objects.filter(lote=lote).order_by('chave_nfe').values(
            'chave_nfe', 'numero_nfe', 'serie_nfe', 'condicao_pagamento_nfe', 'condicao_pagamento_sap'
        )
    )
    if not condicoes:
        return JsonResponse({'sucesso': False, 'mensagem': 'Gere a tabela de condições antes de enviar ao SAP.'}, status=400)
    cod_cliente_sap = getattr(lote.empresa, 'gdfcliente_id', None) if lote.empresa else cod_cliente
    if not cod_cliente_sap:
        cod_cliente_sap = cod_cliente
    try:
        from app.classes.SapRfc import enviar_condicoes_pagamento_sap
        resultado = enviar_condicoes_pagamento_sap(lote.id_lote, cod_cliente_sap, condicoes)
    except Exception as e:
        return JsonResponse({'sucesso': False, 'mensagem': str(e)[:500]}, status=500)
    if not resultado.get('sucesso'):
        return JsonResponse({'sucesso': False, 'mensagem': resultado.get('mensagem', 'Erro ao enviar ao SAP.')}, status=500)
    retornos = resultado.get('retornos') or []
    chaves_lote = {c['chave_nfe'] for c in condicoes}
    atualizados = 0
    for item in retornos:
        chave = item.get('chave_nfe')
        status_sap = item.get('status', 'S')  # U=atualizado, I=processado
        if chave and chave in chaves_lote:
            n = CondicaoPagamentoLote.objects.filter(lote=lote, chave_nfe=chave).update(
                status=status_sap,
            )
            atualizados += n
    return JsonResponse({
        'sucesso': True,
        'mensagem': resultado.get('mensagem', 'Enviado ao SAP.'),
        'enviados': len(retornos),
        'atualizados': atualizados,
    })


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_condicao_param_listar(request):
    """Lista registros da tabela condicao_param (depara condição NFe → SAP) do cliente."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    qs = CondicaoParam.objects.filter(gdfcliente_id=cod_cliente).order_by('condicao_pagamento_nfe', 'tipo_pagamento')
    lista = [
        {
            'id': c.id,
            'condicao_pagamento_nfe': c.condicao_pagamento_nfe or '',
            'condicao_pagamento_sap': c.condicao_pagamento_sap or '',
            'tipo_pagamento': c.tipo_pagamento or '',
        }
        for c in qs
    ]
    return JsonResponse({'sucesso': True, 'condicoes': lista})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["GET"])
def fn_api_reprocessamento_condicao_param_exportar_excel(request):
    """Exporta todos os registros de condicao_param do cliente em planilha .xlsx."""
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return JsonResponse(
            {
                'sucesso': False,
                'mensagem': 'Exportação Excel indisponível: instale o pacote openpyxl (pip install openpyxl).',
            },
            status=503,
        )

    qs = CondicaoParam.objects.filter(gdfcliente_id=cod_cliente).order_by(
        'condicao_pagamento_nfe', 'tipo_pagamento'
    )
    wb = Workbook()
    ws = wb.active
    ws.title = 'Parametros'
    headers = [
        'ID',
        'Tipo pagamento (código)',
        'Tipo pagamento (descrição)',
        'Condição NFe',
        'Condição SAP',
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    for c in qs:
        cod_tipo = (c.tipo_pagamento or '').strip()
        if not cod_tipo:
            tipo_desc = descricao_tipo_pagamento(None)
        else:
            tipo_desc = descricao_tipo_pagamento(cod_tipo) or cod_tipo
        ws.append(
            [
                c.id,
                cod_tipo,
                tipo_desc,
                c.condicao_pagamento_nfe or '',
                c.condicao_pagamento_sap or '',
            ]
        )

    for col_letter, width in zip('ABCDE', [10, 22, 40, 36, 18]):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = timezone.now().strftime('%Y%m%d_%H%M')
    fname = f'parametros_condicao_pagamento_{stamp}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _condicao_param_header_norm(val):
    if val is None:
        return ''
    t = str(val).strip().lower()
    t = ''.join(
        c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn'
    )
    return ' '.join(t.split())


def _condicao_param_parse_id_excel(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return None
        r = round(val)
        if abs(val - r) < 1e-9:
            return int(r)
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(',', '.')))
    except (ValueError, TypeError):
        return None


def _condicao_param_cell_sap(val):
    if val is None:
        return ''
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return ''
        r = round(val)
        if abs(val - r) < 1e-9:
            val = str(int(r))
        else:
            val = str(val).strip()
    else:
        val = str(val).strip()
    return val[:60]


def _condicao_param_row_str(val):
    if val is None:
        return ''
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return ''
        r = round(val)
        if abs(val - r) < 1e-9:
            return str(int(r))
    return str(val).strip()


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_condicao_param_importar_excel(request):
    """
    Carga de planilha .xlsx com o mesmo layout da exportação (aba Parametros).
    Atualiza condicao_pagamento_sap por ID, com checagem de tipo NFe e condição NFe.
    """
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return JsonResponse(
            {
                'sucesso': False,
                'mensagem': 'Importação Excel indisponível: instale o pacote openpyxl (pip install openpyxl).',
            },
            status=503,
        )

    upload = request.FILES.get('arquivo')
    if not upload:
        return JsonResponse({'sucesso': False, 'mensagem': 'Envie um arquivo .xlsx no campo "arquivo".'}, status=400)
    name = (upload.name or '').lower()
    if not name.endswith('.xlsx'):
        return JsonResponse({'sucesso': False, 'mensagem': 'O arquivo deve ser .xlsx.'}, status=400)
    if upload.size > 5 * 1024 * 1024:
        return JsonResponse({'sucesso': False, 'mensagem': 'Arquivo acima do limite de 5 MB.'}, status=400)

    raw = upload.read()
    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return JsonResponse(
            {'sucesso': False, 'mensagem': f'Não foi possível ler a planilha: {str(e)[:200]}'},
            status=400,
        )

    if 'Parametros' in wb.sheetnames:
        ws = wb['Parametros']
    else:
        ws = wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return JsonResponse({'sucesso': False, 'mensagem': 'Planilha vazia.'}, status=400)

    col_map = {}
    for idx, cell in enumerate(header):
        key = _condicao_param_header_norm(cell)
        if key:
            col_map[key] = idx

    key_id = 'id'
    key_sap = 'condicao sap'
    key_tipo = 'tipo pagamento (codigo)'
    key_nfe = 'condicao nfe'

    if key_id not in col_map or key_sap not in col_map:
        wb.close()
        return JsonResponse(
            {
                'sucesso': False,
                'mensagem': 'Cabeçalho inválido: é necessário ID e Condição SAP (use a planilha exportada pelo sistema, sem alterar a linha de títulos).',
            },
            status=400,
        )

    idx_id = col_map[key_id]
    idx_sap = col_map[key_sap]
    idx_tipo = col_map.get(key_tipo)
    idx_nfe = col_map.get(key_nfe)

    atualizados = 0
    linhas_puladas = 0
    detalhes_erro = []
    max_linhas = 5000
    linha_num = 1

    for row in rows:
        linha_num += 1
        if linha_num > max_linhas + 1:
            detalhes_erro.append(f'Processamento interrompido: mais de {max_linhas} linhas de dados.')
            break
        if not row:
            continue
        cells = list(row)
        def cell_at(i):
            if i is None or i >= len(cells):
                return None
            return cells[i]

        pk = _condicao_param_parse_id_excel(cell_at(idx_id))
        sap_new = _condicao_param_cell_sap(cell_at(idx_sap))
        if pk is None:
            if sap_new == '' and all(
                (cell_at(j) is None or str(cell_at(j)).strip() == '')
                for j in (idx_id, idx_sap, idx_tipo, idx_nfe)
                if j is not None
            ):
                continue
            linhas_puladas += 1
            if len(detalhes_erro) < 25:
                detalhes_erro.append(f'Linha {linha_num}: ID inválido ou vazio.')
            continue

        obj = CondicaoParam.objects.filter(pk=pk, gdfcliente_id=cod_cliente).first()
        if not obj:
            linhas_puladas += 1
            if len(detalhes_erro) < 25:
                detalhes_erro.append(f'Linha {linha_num}: ID {pk} não encontrado para este cliente.')
            continue

        if idx_tipo is not None and idx_nfe is not None:
            tipo_x = _condicao_param_row_str(cell_at(idx_tipo))
            nfe_x = _condicao_param_row_str(cell_at(idx_nfe))
            tipo_db = (obj.tipo_pagamento or '').strip()
            nfe_db = (obj.condicao_pagamento_nfe or '').strip()
            if tipo_x != tipo_db or nfe_x != nfe_db:
                linhas_puladas += 1
                if len(detalhes_erro) < 25:
                    detalhes_erro.append(
                        f'Linha {linha_num}: ID {pk} não confere com NFe/tipo do cadastro (não alterado).'
                    )
                continue

        n = CondicaoParam.objects.filter(pk=pk, gdfcliente_id=cod_cliente).update(
            condicao_pagamento_sap=sap_new
        )
        if n:
            atualizados += 1

    wb.close()

    msg_parts = [f'{atualizados} registro(s) atualizado(s).']
    if linhas_puladas:
        msg_parts.append(f'{linhas_puladas} linha(s) ignorada(s).')
    return JsonResponse(
        {
            'sucesso': True,
            'atualizados': atualizados,
            'linhas_ignoradas': linhas_puladas,
            'detalhes_erro': detalhes_erro,
            'mensagem': ' '.join(msg_parts),
        }
    )


@login_required(login_url='Login')
@requer_acesso_subsolucao('Reproc_Painel', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_reprocessamento_condicao_param_atualizar(request):
    """
    Atualiza condicao_pagamento_sap de registros em condicao_param.
    Body: { "itens": [ { "id": 1, "condicao_pagamento_sap": "Z001" }, ... ] }
    """
    cod_cliente = request.session.get('cod_cliente')
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)
    data = json.loads(request.body) if request.body else {}
    itens = data.get('itens') or []
    if not itens:
        return JsonResponse({'sucesso': False, 'mensagem': 'Envie "itens" com id e condicao_pagamento_sap.'}, status=400)
    atualizados = 0
    for item in itens:
        pk = item.get('id')
        if pk is None:
            continue
        cond_sap = (item.get('condicao_pagamento_sap') or '').strip()[:60]
        n = CondicaoParam.objects.filter(pk=pk, gdfcliente_id=cod_cliente).update(condicao_pagamento_sap=cond_sap)
        atualizados += n
    return JsonResponse({'sucesso': True, 'atualizados': atualizados, 'mensagem': f'{atualizados} registro(s) atualizado(s).'})


@login_required(login_url='Login')
@requer_acesso_subsolucao('Dm_Clientes', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_sap_testar_conexao(request):
    """
    Testa a conexão SAP do cliente.
    Body (JSON): { "cod_cliente": "COD" } - opcional; se omitido, usa cod_cliente da sessão.
    Usuário com acesso total pode testar qualquer cliente.
    """
    cod_cliente_sessao = request.session.get('cod_cliente')
    if not cod_cliente_sessao and not usuario_acesso_total_painel(request):
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)

    data = json.loads(request.body) if request.body else {}
    cod_cliente = (data.get('cod_cliente') or '').strip() or cod_cliente_sessao
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Informe cod_cliente ou selecione um cliente na sessão.'}, status=400)
    if not usuario_acesso_total_painel(request) and str(cod_cliente) != str(cod_cliente_sessao):
        return JsonResponse({'sucesso': False, 'mensagem': 'Acesso negado.'}, status=403)

    try:
        from app.classes.SapRfc import SapRfc
        if not SapRfc.is_available():
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'PyRFC não disponível. Instale o SAP NetWeaver RFC SDK e o pacote pyrfc. Ver documentacao_md/',
            }, status=503)
        conn = SapRfc.get_connection(cod_cliente)
        if not conn:
            return JsonResponse({
                'sucesso': False,
                'mensagem': f'Nenhuma conexão SAP ativa para o cliente "{cod_cliente}". Configure na aba Conexão SAP do cliente.',
            }, status=404)
        # Testa conexão chamando RFC de ping (RFC_PING é padrão SAP)
        success, result = SapRfc.call(cod_cliente, 'RFC_PING')
        if success:
            return JsonResponse({
                'sucesso': True,
                'mensagem': 'Conexão SAP OK.',
                'cliente': cod_cliente,
            })
        return JsonResponse({
            'sucesso': False,
            'mensagem': result or 'Falha ao conectar ao SAP.',
            'cliente': cod_cliente,
        }, status=502)
    except Exception as e:
        return JsonResponse({'sucesso': False, 'mensagem': str(e)[:500]}, status=500)


# -------------------------------------------------------------------------
# Integração SAP – RFC (subsolução Int_Rfc)
# -------------------------------------------------------------------------
@login_required(login_url='Login')
@ensure_csrf_cookie
@requer_acesso_subsolucao('Int_Rfc')
@require_http_methods(["GET"])
def fn_view_Integracao_Rfc(request):
    """View da subsolução Int_Rfc: executa RFCs que alimentam tabelas do schema sap."""
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return render(request, 'comum/login.html', {'error_message': 'Cliente não identificado'})

    from app.integracao_sap import get_rfc_registry
    registry = get_rfc_registry()
    rfc_handlers = registry.list_all()

    empresas = list(
        Empresa.objects.filter(
            gdfcliente__cod_cliente=cod_cliente,
            usuarioempresa__user=request.user,
        ).values('cod_empresa', 'fantasia', 'razao').distinct()
    ) if cod_cliente else []
    if usuario_acesso_total_painel(request):
        empresas = list(
            Empresa.objects.filter(gdfcliente__cod_cliente=cod_cliente)
            .values('cod_empresa', 'fantasia', 'razao').distinct()
        ) if cod_cliente else []

    filiais_por_empresa = {}
    for emp in empresas:
        cod_emp = emp.get('cod_empresa')
        filiais = list(
            Filial.objects.filter(empresa__cod_empresa=cod_emp)
            .values('cod_filial', 'nome')
        )
        filiais_por_empresa[cod_emp] = filiais

    url_prefix = (request.META.get("SCRIPT_NAME") or getattr(settings, "FORCE_SCRIPT_NAME", "") or "").strip()
    if url_prefix and not url_prefix.startswith("/"):
        url_prefix = "/" + url_prefix
    url_prefix = url_prefix.rstrip("/")  # '' ou '/gdf'

    context = {
        'cod_cliente': cod_cliente,
        'url_prefix': url_prefix,
        'rfc_handlers': [
            {
                'codigo': h.codigo,
                'nome': h.nome,
                'descricao': h.descricao,
                'tabela_sap': h.tabela_sap,
                'params': [
                    {
                        'key': p.key,
                        'label': p.label,
                        'param_type': p.param_type.value,
                        'required': p.required,
                        'default': p.default,
                        'help_text': p.help_text,
                    }
                    for p in h.params
                ],
            }
            for h in rfc_handlers
        ],
        'empresas': empresas,
        'filiais_por_empresa': filiais_por_empresa,
    }
    return render(request, 'Ferramentas/rfc.html', context)

@login_required(login_url='Login')
@ensure_csrf_cookie 
@requer_acesso_subsolucao('Int_Rfc', redirect_on_deny=False)
@require_http_methods(["POST"])
def fn_api_rfc_executar(request):
    """
    Executa um RFC registrado.
    Body (JSON): { "cod_rfc": "RFC_...", "params": { ... } } — ex.: RFC_RELATORIO_CUSTO, RFC_GDF_RFC_CONSULTA (chaves multilinha).

    Nota: não usar @login_required aqui — ele redireciona para página HTML de login e o fetch
    quebra com JSON.parse (Unexpected token '<'). Autenticação fica em requer_acesso_subsolucao (JSON 403).
    """
    cod_cliente = request.session.get('cod_cliente', None)
    if not cod_cliente:
        return JsonResponse({'sucesso': False, 'mensagem': 'Cliente não identificado'}, status=403)

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'mensagem': 'JSON inválido'}, status=400)

    cod_rfc = (data.get('cod_rfc') or '').strip()
    params = data.get('params') or {}
    if not cod_rfc:
        return JsonResponse({'sucesso': False, 'mensagem': 'cod_rfc é obrigatório'}, status=400)

    from datetime import date as date_type
    from decimal import Decimal
    import uuid

    def _sanitize_for_json(obj):
        if isinstance(obj, dict):
            return {str(k): _sanitize_for_json(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_sanitize_for_json(v) for v in obj]
        if isinstance(obj, (datetime, date_type)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return str(obj)
        if isinstance(obj, uuid.UUID):
            return str(obj)
        if isinstance(obj, bytes):
            return obj.decode('utf-8', errors='replace')
        if isinstance(obj, bool) or obj is None:
            return obj
        if isinstance(obj, (int, str)):
            return obj
        if isinstance(obj, float):
            if obj != obj:  # NaN
                return None
            return obj
        return str(obj)

    from app.integracao_sap import get_rfc_registry
    import logging

    registry = get_rfc_registry()
    try:
        result = registry.execute(cod_rfc, cod_cliente, **params)
    except Exception as e:
        logging.getLogger("gdf").exception(
            "fn_api_rfc_executar cod_rfc=%s cod_cliente=%s", cod_rfc, cod_cliente
        )
        return JsonResponse(
            {
                "sucesso": False,
                "mensagem": f"Erro ao executar RFC: {str(e)[:1500]}",
            },
            status=500,
        )
    try:
        safe = _sanitize_for_json(result)
        return JsonResponse(safe, json_dumps_params={'ensure_ascii': False})
    except (TypeError, ValueError) as e:
        return JsonResponse(
            {
                'sucesso': False,
                'mensagem': f'Erro ao serializar resposta da RFC: {e}',
            },
            status=500,
        )


@csrf_exempt
@require_http_methods(["POST"])
def fn_api_sap_demonstrativos_contabeis(request):
    """
    Consulta RFC /PRCIT/GDF_RFC_BALANCE (demonstrativos contábeis) no processo Django (PyRFC).

    Autenticação: sessão (navegador) ou ``Authorization: Bearer <JWT do dashboard>`` (Streamlit).
    Body JSON: ``i_bukrs``, ``i_ktopl``, ``i_versn``, ``i_year`` e ``i_month_b`` / ``i_month_v``
    (SAP: I_MONTH_B / I_MONTH_V; alias ``i_month_ini`` / ``i_month_fim``), ou ``i_month`` + ``i_year`` (B = V).

    Resposta: ``r_return``, ``arvore`` (nós com ``id``, ``conta``, ``text``, ``valor``, ``children``),
    ``total_nos``, ``periodo``, ``opcoes_arvore``.
    Lista plana: ``parent_id`` na raiz/filhos; campo ``conta`` (conta razão SAP) preservado em cada nó.
    """
    _user, cod_cliente, auth_err = autenticar_sessao_ou_jwt_dashboard(request, "Db_DemonstrContabeis")
    if auth_err is not None:
        return auth_err

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'mensagem': 'JSON inválido'}, status=400)

    from app.integracao_sap.demonstrativos_contabeis_sap import executar_demonstrativos_contabeis

    out = executar_demonstrativos_contabeis(cod_cliente, **data)
    return JsonResponse(out)


# -------------------------------------------------------------------------
# API SAP – Receber Relatório de Custo (POST do SAP → PostgreSQL)
# -------------------------------------------------------------------------
@csrf_exempt
@require_http_methods(["POST"])
def fn_api_sap_relatorio_custo_receber(request):
    """
    Recebe dados de Relatório de Custo enviados pelo SAP via POST e persiste em sap.relatorio_custo.

    Autenticação: Header X-API-Key ou Authorization: Bearer <chave>
    Chave configurada em SAP_RELATORIO_CUSTO_API_KEY (.env).

    Body (JSON):
      {
        "cod_empresa": "1000",      // obrigatório
        "cod_filial": "001",        // opcional
        "registros": [
          { "DOCNUM": "...", "MJAHR": "...", "MBLNR": "...", "PSTDAT": "2025-01-15", ... },
          ...
        ]
      }
    """
    api_key = settings.SAP_RELATORIO_CUSTO_API_KEY
    if not api_key:
        return JsonResponse({
            'sucesso': False,
            'mensagem': 'API não configurada. Defina SAP_RELATORIO_CUSTO_API_KEY no .env.',
        }, status=503)

    # Validar API key
    auth_header = request.headers.get('Authorization', '')
    key_from_header = request.headers.get('X-API-Key', '').strip()
    if not key_from_header and auth_header.startswith('Bearer '):
        key_from_header = auth_header[7:].strip()
    if key_from_header != api_key:
        return JsonResponse({
            'sucesso': False,
            'mensagem': 'API key inválida ou não informada. Use header X-API-Key ou Authorization: Bearer.',
        }, status=401)

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError as e:
        return JsonResponse({
            'sucesso': False,
            'mensagem': f'JSON inválido: {e}',
        }, status=400)

    cod_empresa = (data.get('cod_empresa') or '').strip()
    cod_filial = (data.get('cod_filial') or '').strip() or None
    registros = data.get('registros')

    if registros is None:
        return JsonResponse({
            'sucesso': False,
            'mensagem': 'Campo "registros" é obrigatório (array de objetos).',
        }, status=400)

    if not isinstance(registros, list):
        return JsonResponse({
            'sucesso': False,
            'mensagem': '"registros" deve ser um array.',
        }, status=400)

    from app.integracao_sap.relatorio_custo_receiver import persistir_relatorio_custo
    result = persistir_relatorio_custo(
        cod_empresa=cod_empresa,
        registros=registros,
        cod_filial=cod_filial,
    )
    status_code = 200 if result.get('sucesso') else 400
    return JsonResponse(result, status=status_code)